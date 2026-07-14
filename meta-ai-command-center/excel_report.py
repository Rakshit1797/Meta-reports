#!/usr/bin/env python3
"""
Excel workbook generator for the Meta Ads Performance Intelligence report.

Builds meta_ads_performance_intelligence.xlsx -- a senior-management
performance pack -- from three already-generated, already-validated inputs:

    data/meta_campaign_daily.csv       -- Layer 1 raw campaign daily data
    data/performance_findings.json     -- deterministic analysis (Layer 2)
    reports/meta_performance_report.md -- Claude's executive markdown brief

This script performs presentation-only aggregation (collected-period,
monthly, weekly, and daily rollups; per-campaign collected-period totals)
but reuses performance_analyzer.py's own aggregate_window()/safe_divide()/
safe_pct_change()/funnel/decision helpers rather than reimplementing any
metric definition or threshold -- every ratio and every decision label in
this workbook traces back to that single deterministic module.

The workbook contains exactly seven worksheets, in this order:
    1. Executive Cockpit
    2. Lifetime Performance      (on-sheet title: "Collected Period Performance")
    3. Marketing Funnel
    4. Campaign Portfolio
    5. Trend & Efficiency
    6. AI Management Brief
    7. Raw Data

Website Landings (LPV) is read exclusively from Meta's own
`landing_page_views` column (via performance_analyzer.aggregate_window()).
Clicks are never substituted for it -- if LPV is unavailable, every
LPV-derived cell/chart in this workbook shows "N/A" and is skipped, never
zero.
"""

import argparse
import json
import os
import re
import sys
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from currency_utils import format_currency_text, get_currency_excel_number_format
from performance_analyzer import (
    add_moving_averages,
    aggregate_daily_metrics,
    aggregate_monthly_totals,
    aggregate_weekly_totals,
    aggregate_window,
    classify_portfolio_decision,
    compute_biggest_funnel_leak,
    compute_campaign_shares,
    compute_funnel_stages,
    compute_wasted_spend_candidates,
    compute_top_sales_contributors,
    get_collected_period,
    safe_divide,
    safe_pct_change,
)

DEFAULT_CSV_PATH = os.path.join("data", "meta_campaign_daily.csv")
DEFAULT_FINDINGS_PATH = os.path.join("data", "performance_findings.json")
DEFAULT_REPORT_PATH = os.path.join("reports", "meta_performance_report.md")
DEFAULT_OUTPUT_PATH = "meta_ads_performance_intelligence.xlsx"

REQUIRED_WORKSHEET_ORDER = [
    "Executive Cockpit",
    "Lifetime Performance",
    "Marketing Funnel",
    "Campaign Portfolio",
    "Trend & Efficiency",
    "AI Management Brief",
    "Raw Data",
]
COLLECTED_PERIOD_SHEET_TITLE = "Collected Period Performance"

# ---------------------------------------------------------------------------
# Design system: restrained, board-ready. One header treatment, one positive,
# one warning, one critical, one neutral status color -- nothing else.
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F2937")   # dark charcoal/navy
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=17, color="1F2937")
SUBTITLE_FONT = Font(italic=True, size=10, color="595959")
NOTE_FONT = Font(italic=True, size=9, color="808080")
SECTION_FONT = Font(bold=True, size=13, color="FFFFFF")

POSITIVE_FILL = PatternFill("solid", fgColor="C6EFCE")
WARNING_FILL = PatternFill("solid", fgColor="FFEB9C")
CRITICAL_FILL = PatternFill("solid", fgColor="FFC7CE")
NEUTRAL_FILL = PatternFill("solid", fgColor="D9D9D9")

LINK_FONT = Font(color="0563C1", underline="single", bold=True)

CHANGE_NUMBER_FORMAT = '+0.0"%";-0.0"%";0.0"%"'
CALC_COL_OFFSET = 14  # chart source-data blocks live from column N onward

AI_STATUS_FILLS = {
    "TRACKING WARNING": CRITICAL_FILL,
    "EFFICIENCY RISK": CRITICAL_FILL,
    "CREATIVE FATIGUE": WARNING_FILL,
    "SCALE": POSITIVE_FILL,
    "MONITOR": NEUTRAL_FILL,
    "INSUFFICIENT DATA": NEUTRAL_FILL,
}
PORTFOLIO_DECISION_FILLS = {
    "SCALE": POSITIVE_FILL,
    "PROTECT": POSITIVE_FILL,
    "MAINTAIN": NEUTRAL_FILL,
    "FIX": WARNING_FILL,
    "CUT": CRITICAL_FILL,
    "INVESTIGATE TRACKING": CRITICAL_FILL,
    "INSUFFICIENT DATA": NEUTRAL_FILL,
}

# Columns exactly as produced by meta_collector.py's OUTPUT_COLUMNS. This
# script does not assume any other schema for the raw CSV.
RAW_DATA_COLUMNS = [
    "date", "campaign_id", "campaign_name", "objective", "spend", "impressions", "reach",
    "clicks", "ctr", "cpc", "cpm", "frequency", "purchases", "purchase_value", "CPA", "ROAS",
    "add_to_carts", "initiate_checkouts", "landing_page_views",
    "click_to_landing_page_rate", "landing_page_to_purchase_rate",
]
RAW_DATA_TEXT_COLUMNS = {"campaign_id", "campaign_name", "objective"}
RAW_DATA_DATE_COLUMNS = {"date"}

# Meta's raw `ctr` field is already a percent-number (e.g. 2.35 meaning
# 2.35%), not a 0-1 fraction -- use a custom format that doesn't re-multiply
# by 100. click_to_landing_page_rate / landing_page_to_purchase_rate are
# true 0-1 fractions, so they use Excel's native percentage format.
#
# Currency columns (spend, purchase_value, cpc, cpm, CPA) are NOT listed
# here -- their number_format is built dynamically per the account's actual
# currency via get_currency_excel_number_format(). This workbook never
# hardcodes "$"/USD.
RAW_DATA_NUMBER_FORMATS = {
    "impressions": '#,##0',
    "reach": '#,##0',
    "clicks": '#,##0',
    "purchases": '#,##0',
    "add_to_carts": '#,##0',
    "initiate_checkouts": '#,##0',
    "landing_page_views": '#,##0',
    "frequency": '0.00',
    "ROAS": '0.00',
    "ctr": '0.00"%"',
    "click_to_landing_page_rate": '0.0%',
    "landing_page_to_purchase_rate": '0.0%',
}
RAW_DATA_CURRENCY_COLUMNS = {"spend", "purchase_value", "cpc", "cpm", "CPA"}


class ExcelReportError(Exception):
    """Raised when the Excel workbook cannot be generated reliably."""


# ---------------------------------------------------------------------------
# Loading inputs
# ---------------------------------------------------------------------------

def load_raw_data(csv_path: str) -> pd.DataFrame:
    try:
        df = pd.read_csv(csv_path, dtype={"campaign_id": str})
    except FileNotFoundError as exc:
        raise ExcelReportError(
            f"CSV file '{csv_path}' not found. Run meta_collector.py first."
        ) from exc

    missing_columns = [col for col in RAW_DATA_COLUMNS if col not in df.columns]
    if missing_columns:
        raise ExcelReportError(
            f"CSV file '{csv_path}' is missing expected column(s): {', '.join(missing_columns)}."
        )
    if df.empty:
        raise ExcelReportError(f"CSV file '{csv_path}' contains no rows.")

    df["date"] = pd.to_datetime(df["date"]).dt.date
    numeric_columns = [c for c in RAW_DATA_COLUMNS if c not in RAW_DATA_TEXT_COLUMNS | RAW_DATA_DATE_COLUMNS]
    for col in numeric_columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def load_findings(path: str) -> dict:
    try:
        with open(path) as findings_file:
            return json.load(findings_file)
    except FileNotFoundError as exc:
        raise ExcelReportError(
            f"Findings file '{path}' not found. Run performance_analyzer.py first."
        ) from exc
    except json.JSONDecodeError as exc:
        raise ExcelReportError(f"Findings file '{path}' is not valid JSON: {exc}") from exc


def load_ai_report(path: str) -> str:
    try:
        with open(path) as report_file:
            return report_file.read()
    except FileNotFoundError as exc:
        raise ExcelReportError(
            f"AI report file '{path}' not found. Run ai_analyst.py first."
        ) from exc


# ---------------------------------------------------------------------------
# Aggregation (reuses performance_analyzer.py's math -- no reimplementation)
# ---------------------------------------------------------------------------

def aggregate_campaign_totals(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """Full-collected-period per-campaign rollup, used by Campaign Portfolio."""
    min_date = df["date"].min()
    max_date = df["date"].max()
    campaigns = df[["campaign_id", "campaign_name"]].drop_duplicates()

    results = []
    for _, row in campaigns.iterrows():
        campaign_id = row["campaign_id"]
        campaign_name = row["campaign_name"]
        campaign_rows = df[df["campaign_id"] == campaign_id]

        totals = aggregate_window(campaign_rows, min_date, max_date)

        objective_series = campaign_rows["objective"].dropna()
        objective = str(objective_series.iloc[0]) if not objective_series.empty else ""

        # Reach is not additive across days (it's per-day unique reach), so a
        # true collected-period Frequency (impressions / unique reach) cannot
        # be derived from summed daily reach. The arithmetic mean of the
        # daily Frequency values is used as a pragmatic estimate instead.
        frequency_series = pd.to_numeric(campaign_rows["frequency"], errors="coerce").dropna()
        frequency = float(frequency_series.mean()) if not frequency_series.empty else None

        results.append(
            {
                "campaign_id": campaign_id,
                "campaign_name": campaign_name,
                "objective": objective,
                "frequency": frequency,
                **totals,
            }
        )
    return results


# ---------------------------------------------------------------------------
# Styling / navigation / layout helpers
# ---------------------------------------------------------------------------

def style_header_row(ws, row_idx: int, num_cols: int, start_col: int = 1) -> None:
    for col in range(start_col, start_col + num_cols):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize_columns(ws, max_width: int = 32, min_width: int = 8) -> None:
    # Iterate by column index rather than cell.column_letter -- MergedCell
    # objects (non-anchor cells inside a merged range) don't expose
    # column_letter, but get_column_letter(cell.column) works for any cell.
    for col_cells in ws.columns:
        if not col_cells:
            continue
        col_letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


def hide_gridlines(ws) -> None:
    ws.sheet_view.showGridLines = False


def apply_print_settings(ws, landscape: bool = True) -> None:
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def add_sheet_title(ws, title: str, subtitle: str = "", end_column: int = 8) -> int:
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    return 4


def add_nav_bar(ws, current_sheet: str, row: int = 3) -> None:
    """A single-row navigation strip linking to every other sheet."""
    ws.cell(row=row, column=1, value="Go to:").font = Font(bold=True, size=9, color="595959")
    col = 2
    for sheet_name in REQUIRED_WORKSHEET_ORDER:
        if sheet_name == current_sheet:
            continue
        cell = ws.cell(row=row, column=col, value=sheet_name)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = Font(color="0563C1", underline="single", size=9)
        col += 1


def add_back_link(ws, row: int = 1, column: int = 10) -> None:
    cell = ws.cell(row=row, column=column, value="<< Back to Executive Cockpit")
    cell.hyperlink = "#'Executive Cockpit'!A1"
    cell.font = LINK_FONT


def limit_words(text: str, max_words: int = 100) -> str:
    words = text.split()
    if len(words) <= max_words:
        return text
    return " ".join(words[:max_words]) + "…"


# ---------------------------------------------------------------------------
# Chart helpers -- every chart is sourced from real worksheet cells written
# into a dedicated "Chart Source Data" block, never from inline Python data.
# ---------------------------------------------------------------------------

def _add_line_chart(ws, title, y_title, cats_ref, data_ref, anchor, width=24, height=10, single_series=False):
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = "Date"
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = width
    chart.height = height
    if single_series:
        chart.legend = None
    ws.add_chart(chart, anchor)


def _add_bar_chart(ws, title, x_title, y_title, cats_ref, data_ref, anchor, width=22, height=10, horizontal=False):
    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    chart.title = title
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = width
    chart.height = height
    chart.legend = None
    ws.add_chart(chart, anchor)


def _add_scatter_chart(ws, title, x_title, y_title, x_ref, y_ref, anchor, width=22, height=12):
    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.style = 13
    series = Series(y_ref, x_ref, title=title)
    series.marker = Marker(symbol="circle", size=6)
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    chart.legend = None
    chart.width = width
    chart.height = height
    ws.add_chart(chart, anchor)


# ---------------------------------------------------------------------------
# Sheet 1: Executive Cockpit
# ---------------------------------------------------------------------------

COCKPIT_KPI_DEFINITIONS = [
    ("Collected Period Spend", "spend", "currency", None),
    ("Collected Period Tracked Sales", "purchase_value", "currency", "positive_increase"),
    ("ROAS", "roas", "decimal", "positive_increase"),
    ("Purchases", "purchases", "integer", "positive_increase"),
    ("CPA", "cpa", "currency", "negative_increase"),
    ("Website Landings (LPV)", "landing_page_views", "integer", "positive_increase"),
    ("Landing-to-Purchase CVR", "landing_page_to_purchase_rate", "true_fraction_percent", "positive_increase"),
    ("Reach", "reach", "integer", None),
]


def _format_for_type(format_type: str, currency_code: Optional[str]) -> str:
    if format_type == "currency":
        return get_currency_excel_number_format(currency_code)
    return {
        "integer": '#,##0',
        "decimal": '0.00',
        "true_fraction_percent": '0.0%',
    }[format_type]


def _kpi_status(change: Optional[float], direction: Optional[str], account_confidence: str):
    """Deterministic management status: never treats a raw increase as universally good."""
    if change is None:
        return "N/A", NEUTRAL_FILL
    if account_confidence == "LOW":
        return "WATCH (LOW CONFIDENCE)", WARNING_FILL
    if direction is None:
        return "NEUTRAL", NEUTRAL_FILL
    is_good = (change > 0) if direction == "positive_increase" else (change < 0)
    if change == 0:
        return "FLAT", NEUTRAL_FILL
    return ("ON TRACK", POSITIVE_FILL) if is_good else ("AT RISK", CRITICAL_FILL)


def _direction_arrow(change: Optional[float]) -> str:
    if change is None:
        return "N/A"
    if change > 0:
        return "UP"
    if change < 0:
        return "DOWN"
    return "FLAT"


def build_kpi_cards(ws, last7: dict, previous7: dict, start_row: int, currency_code: Optional[str], account_confidence: str) -> int:
    headers = ["Metric", "Current", "Previous", "% Change", "Direction", "Management Status"]
    for i, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=header)
    style_header_row(ws, start_row, len(headers))

    for offset, (label, metric_key, format_type, direction) in enumerate(COCKPIT_KPI_DEFINITIONS, start=1):
        row = start_row + offset
        fmt = _format_for_type(format_type, currency_code)
        current_value = last7.get(metric_key)
        previous_value = previous7.get(metric_key)
        change = safe_pct_change(current_value, previous_value)

        ws.cell(row=row, column=1, value=label)
        cur_cell = ws.cell(row=row, column=2, value=current_value)
        prev_cell = ws.cell(row=row, column=3, value=previous_value)
        change_cell = ws.cell(row=row, column=4, value=change)
        cur_cell.number_format = fmt
        prev_cell.number_format = fmt
        change_cell.number_format = CHANGE_NUMBER_FORMAT
        ws.cell(row=row, column=5, value=_direction_arrow(change))

        status_label, status_fill = _kpi_status(change, direction, account_confidence)
        status_cell = ws.cell(row=row, column=6, value=status_label)
        status_cell.fill = status_fill

        if current_value is None:
            for c in (1, 2, 3, 4, 5):
                ws.cell(row=row, column=c, value="N/A" if c in (2, 3) else ws.cell(row=row, column=c).value)

    return start_row + len(COCKPIT_KPI_DEFINITIONS)


def build_executive_verdict(ws, start_row: int, verdict_text: str) -> int:
    ws.cell(row=start_row, column=1, value="EXECUTIVE VERDICT").font = SECTION_FONT
    ws.cell(row=start_row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    row = start_row + 1
    cell = ws.cell(row=row, column=1, value=limit_words(verdict_text, 100))
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=6)
    ws.row_dimensions[row].height = 60
    return row + 4


def build_management_signals(ws, start_row: int, signals: List[Dict[str, Any]]) -> int:
    ws.cell(row=start_row, column=1, value="TOP 3 MANAGEMENT SIGNALS").font = SECTION_FONT
    ws.cell(row=start_row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    row = start_row + 1
    headers = ["Signal", "Business Implication", "Confidence"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    style_header_row(ws, row, 1, start_col=1)
    ws.cell(row=row, column=6, value="")
    row += 1
    if not signals:
        ws.cell(row=row, column=1, value="No material evidence identified in the current deterministic findings.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    for signal in signals[:3]:
        ws.cell(row=row, column=1, value=signal["signal"])
        impl_cell = ws.cell(row=row, column=2, value=signal["business_implication"])
        impl_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=6, value=signal["confidence"])
        ws.row_dimensions[row].height = 30
        row += 1
    return row + 1


def build_decisions_required(ws, start_row: int, decisions: List[Dict[str, Any]]) -> int:
    ws.cell(row=start_row, column=1, value="DECISIONS REQUIRED THIS WEEK").font = SECTION_FONT
    ws.cell(row=start_row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    row = start_row + 1
    headers = ["Priority", "Decision", "Evidence", "Commercial Implication", "Confidence"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    if not decisions:
        ws.cell(row=row, column=1, value="No material evidence identified in the current deterministic findings.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
    for decision in decisions[:5]:
        ws.cell(row=row, column=1, value=decision["priority"].upper())
        ws.cell(row=row, column=2, value=decision["decision"]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=decision["evidence"]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=4, value=decision["commercial_implication"]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=5, value=decision["confidence"])
        ws.row_dimensions[row].height = 32
        row += 1
    return row + 1


def build_cockpit_chart_data(ws, start_row: int, daily_totals, funnel_stages, campaign_totals, currency_code) -> Dict[str, int]:
    col = CALC_COL_OFFSET
    row = start_row
    currency_format = get_currency_excel_number_format(currency_code)

    ws.cell(row=row, column=col, value="Chart Source Data (auto-generated -- do not edit)").font = Font(bold=True, size=12)
    row += 2

    daily_header_row = row
    for i, header in enumerate(["Date", "Spend", "Tracked Sales", "ROAS", "CPA", "Purchases"]):
        ws.cell(row=row, column=col + i, value=header)
    style_header_row(ws, row, 6, start_col=col)
    row += 1
    daily_first = row
    for entry in daily_totals:
        ws.cell(row=row, column=col, value=entry["date"]).number_format = "YYYY-MM-DD"
        ws.cell(row=row, column=col + 1, value=entry["spend"]).number_format = currency_format
        ws.cell(row=row, column=col + 2, value=entry["purchase_value"]).number_format = currency_format
        ws.cell(row=row, column=col + 3, value=entry["roas"]).number_format = '0.00'
        ws.cell(row=row, column=col + 4, value=entry["cpa"]).number_format = currency_format
        ws.cell(row=row, column=col + 5, value=entry["purchases"]).number_format = '#,##0'
        row += 1
    daily_last = row - 1
    row += 2

    funnel_header_row = row
    ws.cell(row=row, column=col, value="Stage")
    ws.cell(row=row, column=col + 1, value="Volume")
    style_header_row(ws, row, 2, start_col=col)
    row += 1
    funnel_first = row
    for stage in funnel_stages:
        ws.cell(row=row, column=col, value=stage["stage"])
        vol_cell = ws.cell(row=row, column=col + 1, value=stage["volume"])
        vol_cell.number_format = '#,##0'
        row += 1
    funnel_last = row - 1
    row += 2

    spend_header_row = row
    ws.cell(row=row, column=col, value="Campaign")
    ws.cell(row=row, column=col + 1, value="Spend")
    style_header_row(ws, row, 2, start_col=col)
    row += 1
    spend_first = row
    top_by_spend = sorted(campaign_totals, key=lambda c: c.get("spend") or 0.0, reverse=True)[:10]
    for entry in top_by_spend:
        ws.cell(row=row, column=col, value=entry["campaign_name"])
        ws.cell(row=row, column=col + 1, value=entry["spend"]).number_format = currency_format
        row += 1
    spend_last = row - 1
    row += 2

    sales_header_row = row
    ws.cell(row=row, column=col, value="Campaign")
    ws.cell(row=row, column=col + 1, value="Tracked Sales")
    style_header_row(ws, row, 2, start_col=col)
    row += 1
    sales_first = row
    top_by_sales = sorted(campaign_totals, key=lambda c: c.get("purchase_value") or 0.0, reverse=True)[:10]
    for entry in top_by_sales:
        ws.cell(row=row, column=col, value=entry["campaign_name"])
        ws.cell(row=row, column=col + 1, value=entry["purchase_value"]).number_format = currency_format
        row += 1
    sales_last = row - 1

    return {
        "daily_header_row": daily_header_row, "daily_first": daily_first, "daily_last": daily_last,
        "funnel_header_row": funnel_header_row, "funnel_first": funnel_first, "funnel_last": funnel_last,
        "spend_header_row": spend_header_row, "spend_first": spend_first, "spend_last": spend_last,
        "sales_header_row": sales_header_row, "sales_first": sales_first, "sales_last": sales_last,
    }


def build_executive_cockpit_sheet(wb: Workbook, findings: dict, daily_totals, campaign_totals, ai_sections: Dict[str, str]):
    ws = wb.create_sheet("Executive Cockpit")
    hide_gridlines(ws)
    apply_print_settings(ws)
    currency_code = findings.get("analysis_metadata", {}).get("account_currency")
    account_confidence = findings["account_integrity"]["decision_confidence"]
    meta = findings["analysis_metadata"]

    subtitle = (
        f"Reporting Window: {meta['last_7_day_window']['start']} to {meta['last_7_day_window']['end']}   |   "
        f"Previous Comparable Period: {meta['previous_7_day_window']['start']} to {meta['previous_7_day_window']['end']}   |   "
        f"Account Decision Confidence: {account_confidence}"
    )
    row = add_sheet_title(ws, "Meta Ads Performance Intelligence -- Executive Cockpit", subtitle)
    add_nav_bar(ws, "Executive Cockpit", row=row)
    row += 2

    last7 = findings["account_summary"]["last_7_days"]
    previous7 = findings["account_summary"]["previous_7_days"]
    row = build_kpi_cards(ws, last7, previous7, row, currency_code, account_confidence)
    row += 1

    row = build_executive_verdict(ws, row, ai_sections.get("EXECUTIVE VERDICT", ""))
    row = build_management_signals(ws, row, findings.get("management_signals", []))
    row = build_decisions_required(ws, row, findings.get("management_decisions", []))

    funnel_stages = compute_funnel_stages(last7)
    layout = build_cockpit_chart_data(ws, 1, daily_totals, funnel_stages, campaign_totals, currency_code)

    charts_start_row = row + 2
    spacing = 20
    anchor_row = charts_start_row

    daily_first, daily_last = layout["daily_first"], layout["daily_last"]
    if daily_last >= daily_first:
        cats_ref = Reference(ws, min_col=CALC_COL_OFFSET, min_row=daily_first, max_row=daily_last)
        spend_sales_ref = Reference(
            ws, min_col=CALC_COL_OFFSET + 1, max_col=CALC_COL_OFFSET + 2,
            min_row=layout["daily_header_row"], max_row=daily_last,
        )
        _add_line_chart(ws, "Spend vs Tracked Sales Trend", "Amount", cats_ref, spend_sales_ref, f"A{anchor_row}", width=26, height=11)
        anchor_row += spacing
        roas_ref = Reference(ws, min_col=CALC_COL_OFFSET + 3, max_col=CALC_COL_OFFSET + 3, min_row=layout["daily_header_row"], max_row=daily_last)
        _add_line_chart(ws, "ROAS Trend", "ROAS", cats_ref, roas_ref, f"A{anchor_row}", single_series=True)
        anchor_row += spacing
        cpa_ref = Reference(ws, min_col=CALC_COL_OFFSET + 4, max_col=CALC_COL_OFFSET + 4, min_row=layout["daily_header_row"], max_row=daily_last)
        _add_line_chart(ws, "CPA Trend", "CPA", cats_ref, cpa_ref, f"A{anchor_row}", single_series=True)
        anchor_row += spacing
        purchases_ref = Reference(ws, min_col=CALC_COL_OFFSET + 5, max_col=CALC_COL_OFFSET + 5, min_row=layout["daily_header_row"], max_row=daily_last)
        _add_line_chart(ws, "Purchases Trend", "Purchases", cats_ref, purchases_ref, f"A{anchor_row}", single_series=True)
        anchor_row += spacing

    if layout["funnel_last"] >= layout["funnel_first"]:
        funnel_cats = Reference(ws, min_col=CALC_COL_OFFSET, min_row=layout["funnel_first"], max_row=layout["funnel_last"])
        funnel_data = Reference(ws, min_col=CALC_COL_OFFSET + 1, min_row=layout["funnel_header_row"], max_row=layout["funnel_last"])
        _add_bar_chart(ws, "Marketing Funnel (Last 7 Days)", "Stage", "Count", funnel_cats, funnel_data, f"A{anchor_row}", horizontal=True, height=12)
        anchor_row += spacing

    if layout["spend_last"] >= layout["spend_first"]:
        spend_cats = Reference(ws, min_col=CALC_COL_OFFSET, min_row=layout["spend_first"], max_row=layout["spend_last"])
        spend_data = Reference(ws, min_col=CALC_COL_OFFSET + 1, min_row=layout["spend_header_row"], max_row=layout["spend_last"])
        _add_bar_chart(ws, "Top Campaign Spend Concentration", "Campaign", "Spend", spend_cats, spend_data, f"A{anchor_row}", horizontal=True)
        anchor_row += spacing

    if layout["sales_last"] >= layout["sales_first"]:
        sales_cats = Reference(ws, min_col=CALC_COL_OFFSET, min_row=layout["sales_first"], max_row=layout["sales_last"])
        sales_data = Reference(ws, min_col=CALC_COL_OFFSET + 1, min_row=layout["sales_header_row"], max_row=layout["sales_last"])
        _add_bar_chart(ws, "Top Campaigns by Tracked Sales", "Campaign", "Tracked Sales", sales_cats, sales_data, f"A{anchor_row}", horizontal=True)

    autosize_columns(ws, max_width=30)
    return ws


# ---------------------------------------------------------------------------
# Sheet 2: Lifetime Performance (on-sheet title: "Collected Period Performance")
# ---------------------------------------------------------------------------

COLLECTED_PERIOD_METRICS = [
    ("Total Spend", "spend", "currency"),
    ("Total Tracked Purchase Value", "purchase_value", "currency"),
    ("Total Purchases", "purchases", "integer"),
    ("Total Impressions", "impressions", "integer"),
    ("Total Reach", "reach", "integer"),
    ("Total Clicks", "clicks", "integer"),
    ("Website Landings (LPV)", "landing_page_views", "integer"),
    ("Add to Carts", "add_to_carts", "integer"),
    ("Initiated Checkouts", "initiated_checkouts", "integer"),
    ("ROAS", "roas", "decimal"),
    ("CPA", "cpa", "currency"),
    ("CTR", "ctr", "true_fraction_percent"),
    ("CPC", "cpc", "currency"),
    ("CPM", "cpm", "currency"),
    ("Click-to-Landing Rate", "click_to_landing_page_rate", "true_fraction_percent"),
    ("Landing-to-Purchase CVR", "landing_page_to_purchase_rate", "true_fraction_percent"),
]

MONTHLY_TABLE_COLUMNS = [
    ("Month", "month"), ("Spend", "spend"), ("Tracked Sales", "purchase_value"), ("ROAS", "roas"),
    ("Purchases", "purchases"), ("CPA", "cpa"), ("Impressions", "impressions"), ("Reach", "reach"),
    ("Clicks", "clicks"), ("LPV", "landing_page_views"), ("CTR", "ctr"), ("CPC", "cpc"), ("CPM", "cpm"),
]


def build_collected_period_sheet(wb: Workbook, df: pd.DataFrame, findings: dict, campaign_totals):
    ws = wb.create_sheet("Lifetime Performance")
    hide_gridlines(ws)
    apply_print_settings(ws)
    currency_code = findings.get("analysis_metadata", {}).get("account_currency")
    period = get_collected_period(df)

    subtitle = (
        f"Data available from: {period['min_date'].isoformat()}   |   Data available to: {period['max_date'].isoformat()}"
    )
    row = add_sheet_title(ws, COLLECTED_PERIOD_SHEET_TITLE, subtitle)
    add_nav_bar(ws, "Lifetime Performance", row=row)
    add_back_link(ws, row=1, column=10)
    row += 1

    note = (
        "Collected-period totals represent all data available in this reporting dataset and may "
        "not represent the full lifetime of the Meta ad account."
    )
    ws.cell(row=row, column=1, value=note).font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 2

    account_totals = aggregate_window(df, period["min_date"], period["max_date"])

    ws.cell(row=row, column=1, value="COLLECTED-PERIOD TOTALS").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    totals_start_row = row
    for label, key, fmt_type in COLLECTED_PERIOD_METRICS:
        value = account_totals.get(key)
        ws.cell(row=row, column=1, value=label)
        value_cell = ws.cell(row=row, column=2, value=value if value is not None else "N/A")
        if value is not None:
            value_cell.number_format = _format_for_type(fmt_type, currency_code)
        row += 1
    row += 1

    monthly = aggregate_monthly_totals(df)
    ws.cell(row=row, column=1, value="MONTHLY PERFORMANCE").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(MONTHLY_TABLE_COLUMNS))
    row += 1
    monthly_header_row = row
    for i, (label, _) in enumerate(MONTHLY_TABLE_COLUMNS, start=1):
        ws.cell(row=row, column=i, value=label)
    style_header_row(ws, row, len(MONTHLY_TABLE_COLUMNS))
    row += 1
    monthly_first_row = row
    currency_format = get_currency_excel_number_format(currency_code)
    for entry in monthly:
        for i, (_, key) in enumerate(MONTHLY_TABLE_COLUMNS, start=1):
            value = entry.get(key)
            cell = ws.cell(row=row, column=i, value=value if value is not None else "N/A")
            if key in ("spend", "purchase_value", "cpa", "cpc", "cpm") and value is not None:
                cell.number_format = currency_format
            elif key in ("purchases", "impressions", "reach", "clicks", "landing_page_views") and value is not None:
                cell.number_format = '#,##0'
            elif key == "roas" and value is not None:
                cell.number_format = '0.00'
            elif key == "ctr" and value is not None:
                cell.number_format = '0.00"%"'
        row += 1
    monthly_last_row = row - 1
    row += 2

    # Chart source columns for the monthly charts, sourced from the monthly
    # table itself (real cells, not inline data).
    charts_start_row = row + 2
    spacing = 20
    anchor_row = charts_start_row

    def month_col(key: str) -> int:
        return next(i for i, (_, k) in enumerate(MONTHLY_TABLE_COLUMNS, start=1) if k == key)

    if monthly_last_row >= monthly_first_row:
        cats_ref = Reference(ws, min_col=month_col("month"), min_row=monthly_first_row, max_row=monthly_last_row)
        spend_sales_ref = Reference(
            ws, min_col=month_col("spend"), max_col=month_col("purchase_value"),
            min_row=monthly_header_row, max_row=monthly_last_row,
        )
        _add_line_chart(ws, "Monthly Spend vs Tracked Sales", "Amount", cats_ref, spend_sales_ref, f"A{anchor_row}", width=26, height=11)
        anchor_row += spacing

        roas_ref = Reference(ws, min_col=month_col("roas"), max_col=month_col("roas"), min_row=monthly_header_row, max_row=monthly_last_row)
        _add_line_chart(ws, "Monthly ROAS", "ROAS", cats_ref, roas_ref, f"A{anchor_row}", single_series=True)
        anchor_row += spacing

        cpa_ref = Reference(ws, min_col=month_col("cpa"), max_col=month_col("cpa"), min_row=monthly_header_row, max_row=monthly_last_row)
        _add_line_chart(ws, "Monthly CPA", "CPA", cats_ref, cpa_ref, f"A{anchor_row}", single_series=True)
        anchor_row += spacing

        purchases_ref = Reference(ws, min_col=month_col("purchases"), max_col=month_col("purchases"), min_row=monthly_header_row, max_row=monthly_last_row)
        _add_line_chart(ws, "Monthly Purchases", "Purchases", cats_ref, purchases_ref, f"A{anchor_row}", single_series=True)
        anchor_row += spacing

        # LPV-derived charts are only created when at least one month has a
        # real (non-None) LPV value -- never a fake zero chart.
        if any(entry.get("landing_page_views") is not None for entry in monthly):
            lpv_ref = Reference(ws, min_col=month_col("landing_page_views"), max_col=month_col("landing_page_views"), min_row=monthly_header_row, max_row=monthly_last_row)
            _add_line_chart(ws, "Monthly Website Landings (LPV)", "LPV", cats_ref, lpv_ref, f"A{anchor_row}", single_series=True)
            anchor_row += spacing

        # Landing-to-Purchase CVR isn't a monthly-table column; only add its
        # chart if the underlying rate could be computed from real data.
        cvr_values = [safe_divide(e.get("purchases"), e.get("landing_page_views")) for e in monthly]
        if any(v is not None for v in cvr_values):
            cvr_col = len(MONTHLY_TABLE_COLUMNS) + 1
            ws.cell(row=monthly_header_row, column=cvr_col, value="Landing-to-Purchase CVR")
            style_header_row(ws, monthly_header_row, 1, start_col=cvr_col)
            for offset, v in enumerate(cvr_values):
                cell = ws.cell(row=monthly_first_row + offset, column=cvr_col, value=v)
                cell.number_format = '0.0%'
            cvr_ref = Reference(ws, min_col=cvr_col, max_col=cvr_col, min_row=monthly_header_row, max_row=monthly_last_row)
            _add_line_chart(ws, "Monthly Landing-to-Purchase CVR", "CVR", cats_ref, cvr_ref, f"A{anchor_row}", single_series=True)

    autosize_columns(ws, max_width=26)
    return ws


# ---------------------------------------------------------------------------
# Sheet 3: Marketing Funnel
# ---------------------------------------------------------------------------

def build_marketing_funnel_sheet(wb: Workbook, findings: dict):
    ws = wb.create_sheet("Marketing Funnel")
    hide_gridlines(ws)
    apply_print_settings(ws)

    row = add_sheet_title(ws, "Marketing Funnel -- Impressions to Purchase")
    add_nav_bar(ws, "Marketing Funnel", row=row)
    add_back_link(ws, row=1, column=10)
    row += 1

    current_stages = findings["funnel"]["current_stages"]
    previous_stages = findings["funnel"]["previous_stages"]
    biggest_leak = findings["funnel"]["biggest_leak"]

    headers = [
        "Stage", "Volume", "Stage Conversion Rate", "Previous Period Conversion Rate",
        "Change", "Drop-off Rate", "Management Status", "Formula",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    table_first_row = row
    for cur, prev in zip(current_stages, previous_stages):
        rate = cur["stage_conversion_rate"]
        prev_rate = prev.get("stage_conversion_rate")
        change = safe_pct_change(rate, prev_rate)
        ws.cell(row=row, column=1, value=cur["stage"])
        vol_cell = ws.cell(row=row, column=2, value=cur["volume"] if cur["volume"] is not None else "N/A")
        if cur["volume"] is not None:
            vol_cell.number_format = '#,##0'
        rate_cell = ws.cell(row=row, column=3, value=rate if rate is not None else "N/A")
        if rate is not None:
            rate_cell.number_format = '0.0%'
        prev_rate_cell = ws.cell(row=row, column=4, value=prev_rate if prev_rate is not None else "N/A")
        if prev_rate is not None:
            prev_rate_cell.number_format = '0.0%'
        change_cell = ws.cell(row=row, column=5, value=change if change is not None else "N/A")
        if change is not None:
            change_cell.number_format = CHANGE_NUMBER_FORMAT
        drop_off = cur["drop_off_rate"]
        drop_cell = ws.cell(row=row, column=6, value=drop_off if drop_off is not None else "N/A")
        if drop_off is not None:
            drop_cell.number_format = '0.0%'
            status = "HEALTHY" if drop_off < 0.6 else ("WATCH" if drop_off < 0.85 else "AT RISK")
            fill = POSITIVE_FILL if status == "HEALTHY" else (WARNING_FILL if status == "WATCH" else CRITICAL_FILL)
        else:
            status, fill = "N/A", NEUTRAL_FILL
        status_cell = ws.cell(row=row, column=7, value=status)
        status_cell.fill = fill
        formula_cell = ws.cell(row=row, column=8, value=cur["formula"])
        formula_cell.alignment = Alignment(wrap_text=True)
        row += 1
    table_last_row = row - 1
    row += 2

    cats_ref = Reference(ws, min_col=1, min_row=table_first_row, max_row=table_last_row)
    data_ref = Reference(ws, min_col=2, max_col=2, min_row=table_first_row - 1, max_row=table_last_row)
    _add_bar_chart(ws, "Marketing Funnel -- Collected Period", "Stage", "Volume", cats_ref, data_ref, f"A{row}", horizontal=True, height=14)
    row += 20

    ws.cell(row=row, column=1, value="BIGGEST FUNNEL LEAK").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    leak_headers = ["Stage", "Current Conversion", "Previous Conversion", "Movement", "Commercial Implication"]
    for i, h in enumerate(leak_headers, start=1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(leak_headers))
    row += 1
    if biggest_leak:
        ws.cell(row=row, column=1, value=biggest_leak["stage"])
        cur_cell = ws.cell(row=row, column=2, value=biggest_leak["current_conversion_rate"])
        cur_cell.number_format = '0.0%'
        prev_val = biggest_leak["previous_conversion_rate"]
        prev_cell = ws.cell(row=row, column=3, value=prev_val if prev_val is not None else "N/A")
        if prev_val is not None:
            prev_cell.number_format = '0.0%'
        movement = biggest_leak["movement_pct"]
        move_cell = ws.cell(row=row, column=4, value=movement if movement is not None else "N/A")
        if movement is not None:
            move_cell.number_format = CHANGE_NUMBER_FORMAT
        ws.cell(
            row=row, column=5,
            value=(
                f"{biggest_leak['stage']} is the weakest step in the funnel this period -- "
                "prioritize investigation here before allocating incremental budget."
            ),
        ).alignment = Alignment(wrap_text=True)
    else:
        ws.cell(row=row, column=1, value="No material evidence identified in the current deterministic findings.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2

    ws.cell(row=row, column=1, value="TRACKING INTEGRITY WARNINGS").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    ws.cell(
        row=row, column=1,
        value=(
            "Tracking anomalies below are data-quality signals, not campaign performance "
            "problems -- they must be validated before drawing funnel conclusions."
        ),
    ).font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    warnings = findings.get("account_integrity", {}).get("warnings", [])
    if not warnings:
        ws.cell(row=row, column=1, value="No material evidence identified in the current deterministic findings.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
    for warning in warnings:
        cell = ws.cell(row=row, column=1, value=f"[{warning['severity'].upper()}] {warning['message']}")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 32
        row += 1

    autosize_columns(ws, max_width=32)
    return ws


# ---------------------------------------------------------------------------
# Sheet 4: Campaign Portfolio
# ---------------------------------------------------------------------------

CAMPAIGN_PORTFOLIO_COLUMNS = [
    "Campaign", "Campaign ID", "Spend", "Tracked Sales", "ROAS", "Purchases", "CPA", "Impressions",
    "Reach", "Clicks", "LPV", "CTR", "CPC", "CPM", "Landing-to-Purchase CVR",
    "Spend Share", "Sales Share", "AI Status", "Decision", "Decision Confidence",
]
CAMPAIGN_PORTFOLIO_CURRENCY_COLUMNS = {"Spend", "Tracked Sales", "CPA", "CPC", "CPM"}
CAMPAIGN_PORTFOLIO_INTEGER_COLUMNS = {"Purchases", "Impressions", "Reach", "Clicks", "LPV"}
CAMPAIGN_PORTFOLIO_PERCENT_COLUMNS = {"CTR", "Landing-to-Purchase CVR", "Spend Share", "Sales Share"}


def build_campaign_portfolio_sheet(wb: Workbook, campaign_totals, findings: dict):
    ws = wb.create_sheet("Campaign Portfolio")
    hide_gridlines(ws)
    apply_print_settings(ws)
    currency_code = findings.get("analysis_metadata", {}).get("account_currency")

    row = add_sheet_title(ws, "Campaign Portfolio -- Management Decisions")
    add_nav_bar(ws, "Campaign Portfolio", row=row)
    add_back_link(ws, row=1, column=10)
    row += 2

    confidence_by_campaign = {c["campaign_id"]: c["decision_confidence"] for c in findings.get("campaigns", [])}
    ai_status_by_campaign = {c["campaign_id"]: c["ai_status"] for c in findings.get("campaigns", [])}

    for entry in campaign_totals:
        confidence = confidence_by_campaign.get(entry["campaign_id"], "MEDIUM")
        entry["decision_confidence"] = confidence
        entry["decision"] = classify_portfolio_decision(entry.get("spend"), entry.get("purchases"), entry.get("roas"), confidence)
        entry["ai_status"] = ai_status_by_campaign.get(entry["campaign_id"], "MONITOR")

    for i, header in enumerate(CAMPAIGN_PORTFOLIO_COLUMNS, start=1):
        ws.cell(row=row, column=i, value=header)
    style_header_row(ws, row, len(CAMPAIGN_PORTFOLIO_COLUMNS))
    row += 1
    table_header_row = row - 1
    table_start_row = row

    currency_format = get_currency_excel_number_format(currency_code)
    sorted_totals = sorted(campaign_totals, key=lambda c: c.get("spend") or 0.0, reverse=True)
    for entry in sorted_totals:
        values = [
            entry["campaign_name"], entry["campaign_id"], entry["spend"], entry["purchase_value"],
            entry["roas"], entry["purchases"], entry["cpa"], entry["impressions"], entry["reach"],
            entry["clicks"], entry["landing_page_views"], entry["ctr"], entry["cpc"], entry["cpm"],
            entry["landing_page_to_purchase_rate"], entry["spend_share"], entry["sales_share"],
            entry["ai_status"], entry["decision"], entry["decision_confidence"],
        ]
        ws.append(["N/A" if v is None else v for v in values])
    table_last_row = ws.max_row

    col_index = {name: idx + 1 for idx, name in enumerate(CAMPAIGN_PORTFOLIO_COLUMNS)}
    for r in range(table_start_row, table_last_row + 1):
        ws.cell(row=r, column=col_index["Campaign ID"]).number_format = "@"
    for col_name in CAMPAIGN_PORTFOLIO_CURRENCY_COLUMNS:
        c_idx = col_index[col_name]
        for r in range(table_start_row, table_last_row + 1):
            if isinstance(ws.cell(row=r, column=c_idx).value, (int, float)):
                ws.cell(row=r, column=c_idx).number_format = currency_format
    for col_name in CAMPAIGN_PORTFOLIO_INTEGER_COLUMNS:
        c_idx = col_index[col_name]
        for r in range(table_start_row, table_last_row + 1):
            if isinstance(ws.cell(row=r, column=c_idx).value, (int, float)):
                ws.cell(row=r, column=c_idx).number_format = '#,##0'
    for col_name in CAMPAIGN_PORTFOLIO_PERCENT_COLUMNS:
        c_idx = col_index[col_name]
        for r in range(table_start_row, table_last_row + 1):
            if isinstance(ws.cell(row=r, column=c_idx).value, (int, float)):
                ws.cell(row=r, column=c_idx).number_format = '0.0%'
    roas_idx = col_index["ROAS"]
    for r in range(table_start_row, table_last_row + 1):
        if isinstance(ws.cell(row=r, column=roas_idx).value, (int, float)):
            ws.cell(row=r, column=roas_idx).number_format = '0.00'

    ws.freeze_panes = ws.cell(row=table_start_row, column=1).coordinate
    last_col_letter = get_column_letter(len(CAMPAIGN_PORTFOLIO_COLUMNS))
    table = Table(displayName="CampaignPortfolioTable", ref=f"A{table_header_row}:{last_col_letter}{table_last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    if table_last_row >= table_start_row:
        status_col_letter = get_column_letter(col_index["Decision"])
        status_range = f"{status_col_letter}{table_start_row}:{status_col_letter}{table_last_row}"
        for decision, fill in PORTFOLIO_DECISION_FILLS.items():
            ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=[f'"{decision}"'], fill=fill))
        ai_status_col_letter = get_column_letter(col_index["AI Status"])
        ai_status_range = f"{ai_status_col_letter}{table_start_row}:{ai_status_col_letter}{table_last_row}"
        for status, fill in AI_STATUS_FILLS.items():
            ws.conditional_formatting.add(ai_status_range, CellIsRule(operator="equal", formula=[f'"{status}"'], fill=fill))

    row = table_last_row + 3
    chart_col = CALC_COL_OFFSET
    ws.cell(row=1, column=chart_col, value="Chart Source Data (auto-generated -- do not edit)").font = Font(bold=True, size=12)
    chart_row = 3

    spend_col_idx = col_index["Spend"]
    sales_col_idx = col_index["Tracked Sales"]
    x_ref = Reference(ws, min_col=spend_col_idx, min_row=table_start_row, max_row=table_last_row)
    y_ref = Reference(ws, min_col=sales_col_idx, min_row=table_start_row, max_row=table_last_row)
    anchor_row = row
    if table_last_row >= table_start_row:
        _add_scatter_chart(ws, "Spend vs Tracked Sales by Campaign", "Spend", "Tracked Sales", x_ref, y_ref, f"A{anchor_row}")
        anchor_row += 22

    top10 = sorted_totals[:10]
    if top10:
        top10_start = chart_row + 1
        ws.cell(row=chart_row, column=chart_col, value="Campaign")
        ws.cell(row=chart_row, column=chart_col + 1, value="ROAS")
        ws.cell(row=chart_row, column=chart_col + 2, value="CPA")
        ws.cell(row=chart_row, column=chart_col + 3, value="Spend Share")
        ws.cell(row=chart_row, column=chart_col + 4, value="Sales Share")
        style_header_row(ws, chart_row, 5, start_col=chart_col)
        r = top10_start
        for entry in top10:
            ws.cell(row=r, column=chart_col, value=entry["campaign_name"])
            ws.cell(row=r, column=chart_col + 1, value=entry["roas"] if entry["roas"] is not None else 0).number_format = '0.00'
            ws.cell(row=r, column=chart_col + 2, value=entry["cpa"] if entry["cpa"] is not None else 0).number_format = currency_format
            ws.cell(row=r, column=chart_col + 3, value=entry["spend_share"] if entry["spend_share"] is not None else 0).number_format = '0.0%'
            ws.cell(row=r, column=chart_col + 4, value=entry["sales_share"] if entry["sales_share"] is not None else 0).number_format = '0.0%'
            r += 1
        top10_last = r - 1

        cats_ref = Reference(ws, min_col=chart_col, min_row=top10_start, max_row=top10_last)
        roas_ref = Reference(ws, min_col=chart_col + 1, max_col=chart_col + 1, min_row=chart_row, max_row=top10_last)
        _add_bar_chart(ws, "ROAS by Campaign -- Top 10 by Spend", "Campaign", "ROAS", cats_ref, roas_ref, f"A{anchor_row}", horizontal=True)
        anchor_row += 20
        cpa_ref = Reference(ws, min_col=chart_col + 2, max_col=chart_col + 2, min_row=chart_row, max_row=top10_last)
        _add_bar_chart(ws, "CPA by Campaign -- Top 10 by Spend", "Campaign", "CPA", cats_ref, cpa_ref, f"A{anchor_row}", horizontal=True)
        anchor_row += 20
        share_ref = Reference(ws, min_col=chart_col + 3, max_col=chart_col + 4, min_row=chart_row, max_row=top10_last)
        _add_bar_chart(ws, "Spend Share vs Sales Share", "Campaign", "Share", cats_ref, share_ref, f"A{anchor_row}", horizontal=True)
        anchor_row += 20

    contributors = compute_top_sales_contributors(campaign_totals, max_count=10)
    if contributors:
        contrib_header = chart_row
        contrib_col = chart_col + 6
        ws.cell(row=contrib_header, column=contrib_col, value="Campaign")
        ws.cell(row=contrib_header, column=contrib_col + 1, value="Tracked Sales")
        style_header_row(ws, contrib_header, 2, start_col=contrib_col)
        r = contrib_header + 1
        for entry in contributors:
            ws.cell(row=r, column=contrib_col, value=entry["campaign_name"])
            ws.cell(row=r, column=contrib_col + 1, value=entry["purchase_value"]).number_format = currency_format
            r += 1
        contrib_last = r - 1
        cats_ref = Reference(ws, min_col=contrib_col, min_row=contrib_header + 1, max_row=contrib_last)
        data_ref = Reference(ws, min_col=contrib_col + 1, max_col=contrib_col + 1, min_row=contrib_header, max_row=contrib_last)
        _add_bar_chart(ws, "Top 10 Tracked Sales Contributors", "Campaign", "Tracked Sales", cats_ref, data_ref, f"A{anchor_row}", horizontal=True)
        anchor_row += 20

    wasted = compute_wasted_spend_candidates(campaign_totals)
    if wasted:
        wasted_header = chart_row
        wasted_col = chart_col + 9
        ws.cell(row=wasted_header, column=wasted_col, value="Campaign")
        ws.cell(row=wasted_header, column=wasted_col + 1, value="Wasted Spend")
        style_header_row(ws, wasted_header, 2, start_col=wasted_col)
        r = wasted_header + 1
        for entry in sorted(wasted, key=lambda c: c.get("spend") or 0.0, reverse=True)[:10]:
            ws.cell(row=r, column=wasted_col, value=entry["campaign_name"])
            ws.cell(row=r, column=wasted_col + 1, value=entry["spend"]).number_format = currency_format
            r += 1
        wasted_last = r - 1
        cats_ref = Reference(ws, min_col=wasted_col, min_row=wasted_header + 1, max_row=wasted_last)
        data_ref = Reference(ws, min_col=wasted_col + 1, max_col=wasted_col + 1, min_row=wasted_header, max_row=wasted_last)
        _add_bar_chart(ws, "Top 10 Wasted Spend Campaigns", "Campaign", "Wasted Spend", cats_ref, data_ref, f"A{anchor_row}", horizontal=True)

    autosize_columns(ws, max_width=26)
    return ws


# ---------------------------------------------------------------------------
# Sheet 5: Trend & Efficiency
# ---------------------------------------------------------------------------

DAILY_TABLE_COLUMNS = [
    ("Date", "date"), ("Spend", "spend"), ("Tracked Sales", "purchase_value"), ("ROAS", "roas"),
    ("Purchases", "purchases"), ("CPA", "cpa"), ("CTR", "ctr"), ("CPC", "cpc"), ("CPM", "cpm"),
    ("LPV", "landing_page_views"), ("Click-to-Landing Rate", "click_to_landing_page_rate"),
    ("Landing-to-Purchase CVR", "landing_page_to_purchase_rate"),
]
WEEKLY_TABLE_COLUMNS = [
    ("Week Start", "week_start"), ("Spend", "spend"), ("Tracked Sales", "purchase_value"), ("ROAS", "roas"),
    ("Purchases", "purchases"), ("CPA", "cpa"), ("CTR", "ctr"), ("CPC", "cpc"), ("CPM", "cpm"),
    ("LPV", "landing_page_views"),
]
MOVING_AVERAGE_KEYS = ["roas", "cpa", "ctr", "cpc", "cpm", "landing_page_views", "landing_page_to_purchase_rate"]


def _write_metric_column(ws, row, col, key, value, currency_code):
    currency_format = get_currency_excel_number_format(currency_code)
    if value is None:
        ws.cell(row=row, column=col, value="N/A")
        return
    cell = ws.cell(row=row, column=col, value=value)
    if key in ("spend", "purchase_value", "cpa", "cpc", "cpm"):
        cell.number_format = currency_format
    elif key in ("purchases", "impressions", "reach", "clicks", "landing_page_views"):
        cell.number_format = '#,##0'
    elif key == "roas":
        cell.number_format = '0.00'
    elif key == "ctr":
        cell.number_format = '0.00"%"'
    elif key in ("click_to_landing_page_rate", "landing_page_to_purchase_rate"):
        cell.number_format = '0.0%'


def build_trend_efficiency_sheet(wb: Workbook, df: pd.DataFrame, findings: dict):
    ws = wb.create_sheet("Trend & Efficiency")
    hide_gridlines(ws)
    apply_print_settings(ws)
    currency_code = findings.get("analysis_metadata", {}).get("account_currency")

    row = add_sheet_title(ws, "Trend & Efficiency Analysis")
    add_nav_bar(ws, "Trend & Efficiency", row=row)
    add_back_link(ws, row=1, column=10)
    row += 2

    daily = aggregate_daily_metrics(df)
    add_moving_averages(daily, MOVING_AVERAGE_KEYS)
    weekly = aggregate_weekly_totals(df)

    ws.cell(row=row, column=1, value="DAILY PERFORMANCE").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(DAILY_TABLE_COLUMNS))
    row += 1
    daily_header_row = row
    for i, (label, _) in enumerate(DAILY_TABLE_COLUMNS, start=1):
        ws.cell(row=row, column=i, value=label)
    style_header_row(ws, row, len(DAILY_TABLE_COLUMNS))
    row += 1
    daily_first_row = row
    for entry in daily:
        for i, (_, key) in enumerate(DAILY_TABLE_COLUMNS, start=1):
            if key == "date":
                ws.cell(row=row, column=i, value=entry["date"]).number_format = "YYYY-MM-DD"
            else:
                _write_metric_column(ws, row, i, key, entry.get(key), currency_code)
        row += 1
    daily_last_row = row - 1
    row += 2

    ws.cell(row=row, column=1, value="WEEKLY PERFORMANCE").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(WEEKLY_TABLE_COLUMNS))
    row += 1
    for i, (label, _) in enumerate(WEEKLY_TABLE_COLUMNS, start=1):
        ws.cell(row=row, column=i, value=label)
    style_header_row(ws, row, len(WEEKLY_TABLE_COLUMNS))
    row += 1
    for entry in weekly:
        for i, (_, key) in enumerate(WEEKLY_TABLE_COLUMNS, start=1):
            if key == "week_start":
                ws.cell(row=row, column=i, value=entry["week_start"])
            else:
                _write_metric_column(ws, row, i, key, entry.get(key), currency_code)
        row += 1
    row += 2

    def daily_col(key: str) -> int:
        return next(i for i, (_, k) in enumerate(DAILY_TABLE_COLUMNS, start=1) if k == key)

    anchor_row = row
    spacing = 20
    if daily_last_row >= daily_first_row:
        cats_ref = Reference(ws, min_col=daily_col("date"), min_row=daily_first_row, max_row=daily_last_row)

        spend_sales_ref = Reference(
            ws, min_col=daily_col("spend"), max_col=daily_col("purchase_value"),
            min_row=daily_header_row, max_row=daily_last_row,
        )
        _add_line_chart(ws, "Daily Spend and Tracked Sales", "Amount", cats_ref, spend_sales_ref, f"A{anchor_row}", width=26, height=11)
        anchor_row += spacing

        # 7-day moving averages live in extra columns appended after the
        # visible daily table -- still real worksheet cells, just off to the
        # side, exactly like every other chart source block in this workbook.
        ma_col_start = len(DAILY_TABLE_COLUMNS) + 2
        ws.cell(row=daily_header_row, column=ma_col_start, value="ROAS (7D Avg)")
        ws.cell(row=daily_header_row, column=ma_col_start + 1, value="CPA (7D Avg)")
        style_header_row(ws, daily_header_row, 2, start_col=ma_col_start)
        for offset, entry in enumerate(daily):
            r = daily_first_row + offset
            roas_avg = entry.get("roas_7d_avg")
            cpa_avg = entry.get("cpa_7d_avg")
            ws.cell(row=r, column=ma_col_start, value=roas_avg if roas_avg is not None else None).number_format = '0.00'
            fmt = get_currency_excel_number_format(currency_code)
            cpa_cell = ws.cell(row=r, column=ma_col_start + 1, value=cpa_avg if cpa_avg is not None else None)
            cpa_cell.number_format = fmt

        roas_ma_ref = Reference(ws, min_col=ma_col_start, max_col=ma_col_start, min_row=daily_header_row, max_row=daily_last_row)
        _add_line_chart(ws, "7-Day ROAS Trend", "ROAS (7D Avg)", cats_ref, roas_ma_ref, f"A{anchor_row}", single_series=True)
        anchor_row += spacing
        cpa_ma_ref = Reference(ws, min_col=ma_col_start + 1, max_col=ma_col_start + 1, min_row=daily_header_row, max_row=daily_last_row)
        _add_line_chart(ws, "7-Day CPA Trend", "CPA (7D Avg)", cats_ref, cpa_ma_ref, f"A{anchor_row}", single_series=True)
        anchor_row += spacing

        ctr_ref = Reference(ws, min_col=daily_col("ctr"), max_col=daily_col("ctr"), min_row=daily_header_row, max_row=daily_last_row)
        _add_line_chart(ws, "CTR Trend", "CTR", cats_ref, ctr_ref, f"A{anchor_row}", single_series=True)
        anchor_row += spacing
        cpc_ref = Reference(ws, min_col=daily_col("cpc"), max_col=daily_col("cpc"), min_row=daily_header_row, max_row=daily_last_row)
        _add_line_chart(ws, "CPC Trend", "CPC", cats_ref, cpc_ref, f"A{anchor_row}", single_series=True)
        anchor_row += spacing
        cpm_ref = Reference(ws, min_col=daily_col("cpm"), max_col=daily_col("cpm"), min_row=daily_header_row, max_row=daily_last_row)
        _add_line_chart(ws, "CPM Trend", "CPM", cats_ref, cpm_ref, f"A{anchor_row}", single_series=True)
        anchor_row += spacing

        # LPV-derived trend charts only if LPV was genuinely collected --
        # never plotted as a fake-zero series when unavailable.
        if any(entry.get("landing_page_views") is not None for entry in daily):
            lpv_ref = Reference(ws, min_col=daily_col("landing_page_views"), max_col=daily_col("landing_page_views"), min_row=daily_header_row, max_row=daily_last_row)
            _add_line_chart(ws, "LPV Trend", "LPV", cats_ref, lpv_ref, f"A{anchor_row}", single_series=True)
            anchor_row += spacing
        if any(entry.get("landing_page_to_purchase_rate") is not None for entry in daily):
            cvr_ref = Reference(
                ws, min_col=daily_col("landing_page_to_purchase_rate"), max_col=daily_col("landing_page_to_purchase_rate"),
                min_row=daily_header_row, max_row=daily_last_row,
            )
            _add_line_chart(ws, "Landing-to-Purchase CVR Trend", "CVR", cats_ref, cvr_ref, f"A{anchor_row}", single_series=True)

    autosize_columns(ws, max_width=22)
    return ws


# ---------------------------------------------------------------------------
# Sheet 6: AI Management Brief
# ---------------------------------------------------------------------------

REQUIRED_AI_BRIEF_SECTIONS = [
    "EXECUTIVE VERDICT",
    "WHAT MATERIALLY CHANGED",
    "WHERE MONEY IS BEING WASTED",
    "WHERE INCREMENTAL BUDGET SHOULD MOVE",
    "GROWTH OPPORTUNITIES",
    "FUNNEL & CONVERSION RISKS",
    "TRACKING & MEASUREMENT RISKS",
    "7-DAY ACTION PLAN",
    "MANAGEMENT DECISIONS REQUIRED",
]
AI_BRIEF_SECTION_LABELS = {
    "EXECUTIVE VERDICT": "Executive Verdict",
    "WHAT MATERIALLY CHANGED": "What Materially Changed",
    "WHERE MONEY IS BEING WASTED": "Where Money Is Being Wasted",
    "WHERE INCREMENTAL BUDGET SHOULD MOVE": "Where Incremental Budget Should Move",
    "GROWTH OPPORTUNITIES": "Growth Opportunities",
    "FUNNEL & CONVERSION RISKS": "Funnel & Conversion Risks",
    "TRACKING & MEASUREMENT RISKS": "Tracking & Measurement Risks",
    "7-DAY ACTION PLAN": "7-Day Action Plan",
    "MANAGEMENT DECISIONS REQUIRED": "Management Decisions Required",
}
ACTION_PLAN_SECTIONS = {"7-DAY ACTION PLAN"}
DECISION_SECTIONS = {"MANAGEMENT DECISIONS REQUIRED"}
ACTION_PLAN_FIELDS = ["Priority", "Action", "Campaign", "Evidence", "Expected Business Impact", "Confidence"]
DECISION_FIELDS = ["Priority", "Decision", "Evidence", "Commercial Implication", "Confidence"]
NO_CONTENT_FALLBACK = "No material evidence identified in the current deterministic findings."
LEGACY_FORBIDDEN_PHRASE = "No content generated for this section"

RECOMMENDATION_FIELD_PATTERN = re.compile(
    r'^\**\s*(Priority|Action|Campaign|Evidence|Expected Business Impact|Decision|'
    r'Commercial Implication|Confidence)\s*:\**\s*(.*)$',
    re.IGNORECASE,
)


def parse_ai_report_sections(markdown_text: str) -> Dict[str, str]:
    """Split Claude's markdown into an ordered dict of section name -> body text.

    Matches on '## SECTION NAME' headings (case-insensitive) against the
    required section list. Anything before the first recognized heading is
    treated as part of the Executive Verdict so content is never dropped,
    even if Claude's formatting drifts slightly from the requested structure.
    """
    sections: Dict[str, List[str]] = {name: [] for name in REQUIRED_AI_BRIEF_SECTIONS}
    current = "EXECUTIVE VERDICT"

    for line in markdown_text.splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            heading_text = stripped.lstrip("#").strip().upper()
            matched = next((name for name in REQUIRED_AI_BRIEF_SECTIONS if heading_text == name), None)
            if matched:
                current = matched
                continue
            if stripped.lstrip("#").strip().startswith("Meta Ads"):
                continue
        sections[current].append(line)

    return {name: "\n".join(body_lines).strip() for name, body_lines in sections.items()}


def parse_recommendation_blocks(section_text: str) -> List[Dict[str, str]]:
    """Split one section's body into labeled recommendation/decision blocks.

    Blocks are separated by blank lines. A block that doesn't match the
    expected field structure is kept as free text rather than discarded, so
    no content from Claude is ever lost.
    """
    if not section_text.strip():
        return []

    blocks: List[Dict[str, str]] = []
    for raw_block in re.split(r"\n\s*\n", section_text.strip()):
        fields: Dict[str, str] = {}
        unmatched_lines: List[str] = []
        for line in raw_block.splitlines():
            match = RECOMMENDATION_FIELD_PATTERN.match(line.strip())
            if match:
                field_name = match.group(1).strip()
                canonical = next(
                    (f for f in (ACTION_PLAN_FIELDS + DECISION_FIELDS) if f.lower() == field_name.lower()),
                    field_name,
                )
                fields[canonical] = match.group(2).strip()
            elif line.strip():
                unmatched_lines.append(line.strip())

        if fields:
            if unmatched_lines:
                fields["_free_text"] = " ".join(unmatched_lines)
            blocks.append(fields)
        elif unmatched_lines:
            blocks.append({"free_text": " ".join(unmatched_lines)})

    return blocks


def build_ai_management_brief_sheet(wb: Workbook, markdown_text: str) -> Dict[str, str]:
    ws = wb.create_sheet("AI Management Brief")
    hide_gridlines(ws)
    apply_print_settings(ws, landscape=False)
    sections = parse_ai_report_sections(markdown_text)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95

    row = add_sheet_title(ws, "Meta Ads Performance Intelligence -- AI Management Brief", end_column=2)
    add_nav_bar(ws, "AI Management Brief", row=row)
    add_back_link(ws, row=1, column=4)
    row += 2

    for section_name in REQUIRED_AI_BRIEF_SECTIONS:
        body = sections.get(section_name, "")
        display_label = AI_BRIEF_SECTION_LABELS[section_name]

        header_cell = ws.cell(row=row, column=1, value=display_label)
        header_cell.font = SECTION_FONT
        header_cell.fill = HEADER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 22
        row += 1

        if section_name == "EXECUTIVE VERDICT":
            text = limit_words(body, 100) if body else NO_CONTENT_FALLBACK
            cell = ws.cell(row=row, column=1, value=text)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.row_dimensions[row].height = 60
            row += 2
            continue

        if not body:
            ws.cell(row=row, column=1, value=NO_CONTENT_FALLBACK)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 2
            continue

        if section_name in ACTION_PLAN_SECTIONS or section_name in DECISION_SECTIONS:
            field_list = ACTION_PLAN_FIELDS if section_name in ACTION_PLAN_SECTIONS else DECISION_FIELDS
            blocks = parse_recommendation_blocks(body)
            max_rows = 5
            blocks = blocks[:max_rows] if blocks else blocks
            if not blocks:
                cell = ws.cell(row=row, column=1, value=body)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                ws.row_dimensions[row].height = 60
                row += 2
                continue
            for block in blocks:
                if "free_text" in block:
                    cell = ws.cell(row=row, column=1, value=block["free_text"])
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                    ws.row_dimensions[row].height = 45
                    row += 1
                    continue
                for field_name in field_list:
                    label_cell = ws.cell(row=row, column=1, value=f"{field_name}:")
                    label_cell.font = Font(bold=True)
                    label_cell.alignment = Alignment(vertical="top")
                    value_cell = ws.cell(row=row, column=2, value=block.get(field_name, ""))
                    value_cell.alignment = Alignment(wrap_text=True, vertical="top")
                    ws.row_dimensions[row].height = 32
                    row += 1
                row += 1  # blank separator between blocks
        else:
            cell = ws.cell(row=row, column=1, value=body)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            approx_lines = max(1, body.count("\n") + len(body) // 90)
            ws.row_dimensions[row].height = min(400, max(30, approx_lines * 15))
            row += 2

    ws.freeze_panes = "A1"
    return sections


# ---------------------------------------------------------------------------
# Sheet 7: Raw Data
# ---------------------------------------------------------------------------

def build_raw_data_sheet(wb: Workbook, df: pd.DataFrame, currency_code: Optional[str] = None):
    ws = wb.create_sheet("Raw Data")
    ws.append(RAW_DATA_COLUMNS)
    style_header_row(ws, 1, len(RAW_DATA_COLUMNS))

    for _, row_data in df.iterrows():
        values = []
        for col in RAW_DATA_COLUMNS:
            value = row_data[col]
            if col in RAW_DATA_DATE_COLUMNS:
                values.append(value)
            elif col in RAW_DATA_TEXT_COLUMNS:
                values.append("" if pd.isna(value) else str(value))
            else:
                values.append(None if pd.isna(value) else float(value))
        ws.append(values)

    last_row = ws.max_row
    col_index = {name: idx + 1 for idx, name in enumerate(RAW_DATA_COLUMNS)}

    for r in range(2, last_row + 1):
        ws.cell(row=r, column=col_index["date"]).number_format = "YYYY-MM-DD"
        ws.cell(row=r, column=col_index["campaign_id"]).number_format = "@"

    for col_name, fmt in RAW_DATA_NUMBER_FORMATS.items():
        c_idx = col_index[col_name]
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=c_idx).number_format = fmt

    currency_format = get_currency_excel_number_format(currency_code)
    for col_name in RAW_DATA_CURRENCY_COLUMNS:
        c_idx = col_index[col_name]
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=c_idx).number_format = currency_format

    ws.freeze_panes = "A2"

    last_col_letter = get_column_letter(len(RAW_DATA_COLUMNS))
    table = Table(displayName="RawDataTable", ref=f"A1:{last_col_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    add_back_link(ws, row=1, column=len(RAW_DATA_COLUMNS) + 2)
    autosize_columns(ws, max_width=24)
    return ws


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def generate_workbook(csv_path: str, findings_path: str, report_path: str, output_path: str) -> str:
    df = load_raw_data(csv_path)
    findings = load_findings(findings_path)
    markdown_text = load_ai_report(report_path)

    campaign_totals = aggregate_campaign_totals(df)
    period = get_collected_period(df)
    account_totals = aggregate_window(df, period["min_date"], period["max_date"])
    compute_campaign_shares(campaign_totals, account_totals)

    daily_totals = [
        {
            "date": entry["date"],
            "spend": entry["spend"],
            "purchase_value": entry["purchase_value"],
            "roas": entry["roas"],
            "cpa": entry["cpa"],
            "purchases": entry["purchases"],
        }
        for entry in aggregate_daily_metrics(df)
    ]
    currency_code = findings.get("analysis_metadata", {}).get("account_currency")

    wb = Workbook()
    wb.remove(wb.active)

    # AI Management Brief is built first (off-workbook parse) so its
    # Executive Verdict section can also drive the Executive Cockpit sheet --
    # both sheets read from the exact same parsed AI markdown.
    ai_sections = parse_ai_report_sections(markdown_text)

    build_executive_cockpit_sheet(wb, findings, daily_totals, campaign_totals, ai_sections)
    build_collected_period_sheet(wb, df, findings, campaign_totals)
    build_marketing_funnel_sheet(wb, findings)
    build_campaign_portfolio_sheet(wb, campaign_totals, findings)
    build_trend_efficiency_sheet(wb, df, findings)
    build_ai_management_brief_sheet(wb, markdown_text)
    build_raw_data_sheet(wb, df, currency_code=currency_code)

    if list(wb.sheetnames) != REQUIRED_WORKSHEET_ORDER:
        raise ExcelReportError(
            f"Internal error: worksheet order {list(wb.sheetnames)} does not match "
            f"the required order {REQUIRED_WORKSHEET_ORDER}."
        )

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate the Meta Ads Performance Intelligence Excel workbook.")
    parser.add_argument("--csv", default=DEFAULT_CSV_PATH, help="Path to meta_campaign_daily.csv.")
    parser.add_argument("--findings", default=DEFAULT_FINDINGS_PATH, help="Path to performance_findings.json.")
    parser.add_argument("--report", default=DEFAULT_REPORT_PATH, help="Path to meta_performance_report.md.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_PATH, help="Output .xlsx path.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    try:
        generate_workbook(args.csv, args.findings, args.report, args.output)
    except ExcelReportError as exc:
        print(f"Excel report error: {exc}")
        sys.exit(1)
    print(f"Saved Excel workbook to {args.output}")


if __name__ == "__main__":
    main()
