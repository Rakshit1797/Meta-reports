#!/usr/bin/env python3
"""
Unit tests for excel_report.py.

Builds small synthetic CSV/findings/report fixtures on disk (no live Meta
API or Anthropic API calls) and verifies the generated workbook's structure,
content, and chart wiring. Run with:

    python -m unittest discover -s tests -t . -v
"""

import json
import os
import shutil
import sys
import tempfile
import unittest
from datetime import date, timedelta

import openpyxl
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from excel_report import (  # noqa: E402
    COCKPIT_KPI_DEFINITIONS,
    NO_CONTENT_FALLBACK,
    LEGACY_FORBIDDEN_PHRASE,
    REQUIRED_WORKSHEET_ORDER,
    aggregate_campaign_totals,
    generate_workbook,
    limit_words,
    parse_ai_report_sections,
    parse_recommendation_blocks,
)


def _build_synthetic_csv(path: str, days: int = 42, include_lpv: bool = True) -> None:
    """Two campaigns over `days` days, spanning at least one full month boundary."""
    rows = []
    end = date(2026, 7, 14)
    start = end - timedelta(days=days - 1)

    def add_campaign(campaign_id, name, objective, vals):
        for offset in range(days):
            d = start + timedelta(days=offset)
            rows.append(
                {"date": d.isoformat(), "campaign_id": campaign_id, "campaign_name": name, "objective": objective, **vals}
            )

    healthy = {
        "spend": 100, "impressions": 3000, "reach": 2500, "clicks": 80, "ctr": 2.5, "cpc": 1.25,
        "cpm": 33.3, "frequency": 1.1, "purchases": 5, "purchase_value": 500,
        "add_to_carts": 20, "initiate_checkouts": 10, "landing_page_views": 50,
    }
    zero_purchase = {
        "spend": 300, "impressions": 4000, "reach": 3000, "clicks": 90, "ctr": 2.2, "cpc": 3.3,
        "cpm": 75, "frequency": 1.0, "purchases": 0, "purchase_value": 0,
        "add_to_carts": 5, "initiate_checkouts": 2, "landing_page_views": 40,
    }

    add_campaign("300000000001", "Steady Campaign", "OUTCOME_SALES", healthy)
    add_campaign("300000000002", "Zero Purchase Campaign", "OUTCOME_TRAFFIC", zero_purchase)

    df = pd.DataFrame(rows)

    def safe_div(n, d):
        return (n / d) if d else None

    df["CPA"] = df.apply(lambda r: safe_div(r["spend"], r["purchases"]), axis=1)
    df["ROAS"] = df.apply(lambda r: safe_div(r["purchase_value"], r["spend"]), axis=1)
    df["click_to_landing_page_rate"] = df.apply(lambda r: safe_div(r["landing_page_views"], r["clicks"]), axis=1)
    df["landing_page_to_purchase_rate"] = df.apply(lambda r: safe_div(r["purchases"], r["landing_page_views"]), axis=1)

    if not include_lpv:
        df["landing_page_views"] = None
        df["click_to_landing_page_rate"] = None
        df["landing_page_to_purchase_rate"] = None

    column_order = [
        "date", "campaign_id", "campaign_name", "objective", "spend", "impressions", "reach", "clicks",
        "ctr", "cpc", "cpm", "frequency", "purchases", "purchase_value", "CPA", "ROAS", "add_to_carts",
        "initiate_checkouts", "landing_page_views", "click_to_landing_page_rate", "landing_page_to_purchase_rate",
    ]
    df[column_order].to_csv(path, index=False)


SAMPLE_REPORT_MARKDOWN = """# Meta Ads Performance Intelligence Report

## EXECUTIVE VERDICT
Account performance was mixed over the collected period. Spend increased while efficiency held broadly steady, with one campaign showing zero purchases despite meaningful spend that requires prompt review.

## WHAT MATERIALLY CHANGED
FACT: CPA moved modestly period over period.

## WHERE MONEY IS BEING WASTED
Priority: critical
Action: Pause Zero Purchase Campaign pending review
Campaign: Zero Purchase Campaign
Evidence: Spend of 2100.00 with 0 purchases in the last 7 days.
Expected Business Impact: Stop further wasted spend.
Confidence: HIGH

## WHERE INCREMENTAL BUDGET SHOULD MOVE
No issues detected in this category for the current period.

## GROWTH OPPORTUNITIES
No issues detected in this category for the current period.

## FUNNEL & CONVERSION RISKS
FACT: Funnel metrics were within normal range account-wide.

## TRACKING & MEASUREMENT RISKS
No issues detected in this category for the current period.

## 7-DAY ACTION PLAN

Priority: critical
Action: Pause and review Zero Purchase Campaign targeting
Campaign: Zero Purchase Campaign
Evidence: Spend of 2100.00 with 0 purchases in the last 7 days.
Expected Business Impact: Prevent further wasted spend.
Confidence: HIGH

## MANAGEMENT DECISIONS REQUIRED

Priority: critical
Decision: Pause Zero Purchase Campaign pending targeting review
Evidence: Spend of 2100.00 with 0 purchases in the last 7 days.
Commercial Implication: Avoid continued wasted spend.
Confidence: HIGH
"""

EMPTY_SECTIONS_REPORT_MARKDOWN = """# Meta Ads Performance Intelligence Report

## EXECUTIVE VERDICT


## WHAT MATERIALLY CHANGED


## WHERE MONEY IS BEING WASTED


## WHERE INCREMENTAL BUDGET SHOULD MOVE


## GROWTH OPPORTUNITIES


## FUNNEL & CONVERSION RISKS


## TRACKING & MEASUREMENT RISKS


## 7-DAY ACTION PLAN


## MANAGEMENT DECISIONS REQUIRED

"""


class ExcelReportTestCase(unittest.TestCase):
    """Base class that builds a full synthetic pipeline fixture in a temp dir."""

    CSV_DAYS = 42
    INCLUDE_LPV = True
    REPORT_MARKDOWN = SAMPLE_REPORT_MARKDOWN

    def setUp(self):
        self.tmp_dir = tempfile.mkdtemp()
        self.csv_path = os.path.join(self.tmp_dir, "meta_campaign_daily.csv")
        self.findings_path = os.path.join(self.tmp_dir, "performance_findings.json")
        self.report_path = os.path.join(self.tmp_dir, "meta_performance_report.md")
        self.output_path = os.path.join(self.tmp_dir, "meta_ads_performance_intelligence.xlsx")

        _build_synthetic_csv(self.csv_path, days=self.CSV_DAYS, include_lpv=self.INCLUDE_LPV)

        # Run the real (unmodified) performance_analyzer.py to produce findings --
        # no live API calls are involved, this is pure local computation.
        from performance_analyzer import analyze, load_dataset

        df = load_dataset(self.csv_path)
        results = analyze(df)
        with open(self.findings_path, "w") as f:
            json.dump(results, f, indent=2, default=str)

        with open(self.report_path, "w") as f:
            f.write(self.REPORT_MARKDOWN)

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def set_account_currency(self, currency_code):
        """Rewrite the findings fixture's account_currency, then regenerate the workbook."""
        with open(self.findings_path) as f:
            findings = json.load(f)
        findings["analysis_metadata"]["account_currency"] = currency_code
        with open(self.findings_path, "w") as f:
            json.dump(findings, f, indent=2, default=str)
        generate_workbook(self.csv_path, self.findings_path, self.report_path, self.output_path)
        return openpyxl.load_workbook(self.output_path)

    def generate(self):
        generate_workbook(self.csv_path, self.findings_path, self.report_path, self.output_path)
        return openpyxl.load_workbook(self.output_path)


class TestWorkbookCreation(ExcelReportTestCase):
    def test_workbook_is_created_and_openable(self):
        generate_workbook(self.csv_path, self.findings_path, self.report_path, self.output_path)
        self.assertTrue(os.path.isfile(self.output_path))
        wb = openpyxl.load_workbook(self.output_path)
        self.assertIsNotNone(wb)


class TestWorksheetNamesAndOrder(ExcelReportTestCase):
    def test_exactly_seven_worksheets_in_required_order(self):
        wb = self.generate()
        self.assertEqual(list(wb.sheetnames), REQUIRED_WORKSHEET_ORDER)
        self.assertEqual(len(wb.sheetnames), 7)
        self.assertEqual(
            REQUIRED_WORKSHEET_ORDER,
            [
                "Executive Cockpit", "Lifetime Performance", "Marketing Funnel",
                "Campaign Portfolio", "Trend & Efficiency", "AI Management Brief", "Raw Data",
            ],
        )


class TestExecutiveCockpitKpiCards(ExcelReportTestCase):
    def test_all_required_kpi_cards_present(self):
        wb = self.generate()
        ws = wb["Executive Cockpit"]
        labels_found = {cell.value for row in ws.iter_rows(min_col=1, max_col=1) for cell in row if cell.value}
        for label, _, _, _ in COCKPIT_KPI_DEFINITIONS:
            self.assertIn(label, labels_found)

    def test_kpi_cards_have_numeric_current_and_previous_values(self):
        wb = self.generate()
        ws = wb["Executive Cockpit"]
        numeric_found = any(
            isinstance(cell.value, (int, float))
            for row in ws.iter_rows(min_col=2, max_col=3, max_row=25)
            for cell in row
        )
        self.assertTrue(numeric_found)


class TestCollectedPeriodDisclosure(ExcelReportTestCase):
    def test_data_available_dates_disclosed(self):
        wb = self.generate()
        ws = wb["Lifetime Performance"]
        subtitle_texts = [
            cell.value for row in ws.iter_rows(min_row=1, max_row=3) for cell in row if isinstance(cell.value, str)
        ]
        combined = " ".join(subtitle_texts)
        self.assertIn("Data available from", combined)
        self.assertIn("Data available to", combined)

    def test_sheet_title_reads_collected_period_performance(self):
        wb = self.generate()
        ws = wb["Lifetime Performance"]
        self.assertIn("Collected Period Performance", ws.cell(row=1, column=1).value)

    def test_collected_period_note_present(self):
        wb = self.generate()
        ws = wb["Lifetime Performance"]
        all_text = " ".join(
            cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)
        )
        self.assertIn("may not represent the full lifetime of the Meta ad account", all_text)


class TestLpvAvailable(ExcelReportTestCase):
    INCLUDE_LPV = True

    def test_lpv_shown_as_real_number_in_raw_data(self):
        wb = self.generate()
        ws = wb["Raw Data"]
        header = [c.value for c in ws[1]]
        lpv_col = header.index("landing_page_views") + 1
        value = ws.cell(row=2, column=lpv_col).value
        self.assertIsInstance(value, (int, float))
        self.assertGreater(value, 0)

    def test_lpv_kpi_card_shows_numeric_value(self):
        wb = self.generate()
        ws = wb["Executive Cockpit"]
        row = self._find_row(ws, "Website Landings (LPV)")
        self.assertIsInstance(ws.cell(row=row, column=2).value, (int, float))

    @staticmethod
    def _find_row(ws, label):
        for row in ws.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if cell.value == label:
                    return cell.row
        raise AssertionError(f"Label '{label}' not found")


class TestLpvUnavailable(ExcelReportTestCase):
    INCLUDE_LPV = False

    def test_lpv_kpi_card_shows_na_not_zero(self):
        wb = self.generate()
        ws = wb["Executive Cockpit"]
        row = TestLpvAvailable._find_row(ws, "Website Landings (LPV)")
        current_value = ws.cell(row=row, column=2).value
        self.assertIn(current_value, ("N/A", None))
        self.assertNotEqual(current_value, 0)

    def test_clicks_are_never_substituted_for_missing_lpv(self):
        wb = self.generate()
        ws = wb["Raw Data"]
        header = [c.value for c in ws[1]]
        lpv_col = header.index("landing_page_views") + 1
        clicks_col = header.index("clicks") + 1
        lpv_value = ws.cell(row=2, column=lpv_col).value
        clicks_value = ws.cell(row=2, column=clicks_col).value
        self.assertNotEqual(lpv_value, clicks_value)

    def test_no_lpv_derived_charts_created_when_unavailable(self):
        wb = self.generate()
        ws = wb["Trend & Efficiency"]
        titles = []
        for chart in ws._charts:
            try:
                titles.append(chart.title.tx.rich.p[0].r[0].t)
            except Exception:
                pass
        self.assertNotIn("LPV Trend", titles)
        self.assertNotIn("Landing-to-Purchase CVR Trend", titles)


class TestCurrencyFormatting(ExcelReportTestCase):
    """Covers requirement: Excel currency formatting must dynamically use the
    account currency -- never hardcoded '$', and never assumed to be USD when
    currency metadata is missing."""

    def _raw_data_spend_format(self, wb):
        ws = wb["Raw Data"]
        header = [c.value for c in ws[1]]
        return ws.cell(row=2, column=header.index("spend") + 1).number_format

    def _portfolio_spend_format(self, wb):
        ws = wb["Campaign Portfolio"]
        header = [c.value for c in ws[6]]  # header row (after title/nav/blank rows)
        header = [c.value for row in ws.iter_rows(min_row=1, max_row=10) for c in row if c.value == "Spend"]
        # locate header row robustly
        for row in ws.iter_rows():
            values = [c.value for c in row]
            if "Spend" in values:
                col = values.index("Spend") + 1
                return ws.cell(row=row[0].row + 1, column=col).number_format
        raise AssertionError("Spend column not found in Campaign Portfolio")

    def test_inr_uses_rupee_symbol(self):
        wb = self.set_account_currency("INR")
        self.assertEqual(self._raw_data_spend_format(wb), '"₹"#,##0.00')
        self.assertEqual(self._portfolio_spend_format(wb), '"₹"#,##0.00')

    def test_usd_uses_dollar_sign(self):
        wb = self.set_account_currency("USD")
        self.assertEqual(self._raw_data_spend_format(wb), '"$"#,##0.00')
        self.assertEqual(self._portfolio_spend_format(wb), '"$"#,##0.00')

    def test_missing_currency_falls_back_without_assuming_usd(self):
        wb = self.set_account_currency(None)
        fmt = self._raw_data_spend_format(wb)
        self.assertEqual(fmt, '#,##0.00')
        self.assertNotIn("$", fmt)

    def test_unknown_currency_code_appends_iso_code(self):
        wb = self.set_account_currency("AED")
        fmt = self._raw_data_spend_format(wb)
        self.assertEqual(fmt, '#,##0.00" AED"')
        self.assertNotIn("$", fmt)


class TestMarketingFunnelSheet(ExcelReportTestCase):
    def test_funnel_stage_rows_match_deterministic_stages(self):
        wb = self.generate()
        ws = wb["Marketing Funnel"]
        stage_labels = [
            cell.value for row in ws.iter_rows(min_col=1, max_col=1) for cell in row if isinstance(cell.value, str)
        ]
        for expected in ("Impressions", "Reach", "Clicks", "Website Landings (LPV)", "Add to Cart", "Initiated Checkout", "Purchase"):
            self.assertIn(expected, stage_labels)

    def test_biggest_funnel_leak_section_present(self):
        wb = self.generate()
        ws = wb["Marketing Funnel"]
        all_text = " ".join(
            cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)
        )
        self.assertIn("BIGGEST FUNNEL LEAK", all_text)

    def test_tracking_integrity_warnings_labeled_as_tracking_not_performance(self):
        wb = self.generate()
        ws = wb["Marketing Funnel"]
        all_text = " ".join(
            cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)
        )
        self.assertIn("TRACKING INTEGRITY WARNINGS", all_text)
        self.assertIn("not campaign performance problems", all_text)


class TestCampaignPortfolioSheet(ExcelReportTestCase):
    def _read_table(self, wb):
        ws = wb["Campaign Portfolio"]
        header_row = None
        for row in ws.iter_rows():
            values = [c.value for c in row]
            if "Campaign" in values and "Decision" in values:
                header_row = row[0].row
                header = values
                break
        self.assertIsNotNone(header_row)
        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            values = [ws.cell(row=r, column=c + 1).value for c in range(len(header))]
            if values[0] is None:
                continue
            rows.append(dict(zip(header, values)))
        return rows

    def test_spend_share_present_and_bounded(self):
        wb = self.generate()
        rows = self._read_table(wb)
        for row in rows:
            self.assertIsInstance(row["Spend Share"], (int, float))
            self.assertGreaterEqual(row["Spend Share"], 0.0)
            self.assertLessEqual(row["Spend Share"], 1.0)

    def test_sales_share_present_and_bounded(self):
        wb = self.generate()
        rows = self._read_table(wb)
        for row in rows:
            self.assertIsInstance(row["Sales Share"], (int, float))
            self.assertGreaterEqual(row["Sales Share"], 0.0)
            self.assertLessEqual(row["Sales Share"], 1.0)

    def test_decision_column_uses_allowed_labels_only(self):
        from performance_analyzer import PORTFOLIO_DECISIONS

        wb = self.generate()
        rows = self._read_table(wb)
        for row in rows:
            self.assertIn(row["Decision"], PORTFOLIO_DECISIONS)


class TestMonthlyWeeklyAndMovingAverage(ExcelReportTestCase):
    CSV_DAYS = 45  # spans a calendar-month boundary

    def test_monthly_table_has_multiple_rows(self):
        wb = self.generate()
        ws = wb["Lifetime Performance"]
        month_values = [
            cell.value
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and len(cell.value) == 7 and cell.value[4] == "-"
        ]
        self.assertGreaterEqual(len(set(month_values)), 2)

    def test_weekly_table_present_in_trend_sheet(self):
        wb = self.generate()
        ws = wb["Trend & Efficiency"]
        all_text = [cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)]
        self.assertIn("WEEKLY PERFORMANCE", all_text)
        self.assertIn("Week Start", all_text)

    def test_seven_day_moving_average_column_present(self):
        wb = self.generate()
        ws = wb["Trend & Efficiency"]
        all_text = [cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)]
        self.assertIn("ROAS (7D Avg)", all_text)
        self.assertIn("CPA (7D Avg)", all_text)


class TestRawDataPreservation(ExcelReportTestCase):
    def test_raw_data_row_count_matches_csv_exactly(self):
        wb = self.generate()
        ws = wb["Raw Data"]
        with open(self.csv_path) as csv_file:
            csv_row_count = sum(1 for _ in csv_file) - 1  # minus header
        self.assertEqual(ws.max_row - 1, csv_row_count)

    def test_raw_data_campaign_id_preserved_as_text(self):
        wb = self.generate()
        ws = wb["Raw Data"]
        header = [c.value for c in ws[1]]
        campaign_id_col = header.index("campaign_id") + 1
        value = ws.cell(row=2, column=campaign_id_col).value
        self.assertIsInstance(value, str)
        self.assertEqual(value, "300000000001")

    def test_raw_data_values_not_modified(self):
        wb = self.generate()
        ws = wb["Raw Data"]
        header = [c.value for c in ws[1]]
        spend_col = header.index("spend") + 1
        first_data_spend = ws.cell(row=2, column=spend_col).value
        df = pd.read_csv(self.csv_path)
        self.assertAlmostEqual(float(first_data_spend), float(df["spend"].iloc[0]))


class TestCampaignMetricAggregation(unittest.TestCase):
    def test_ctr_computed_from_totals_not_averaged(self):
        df = pd.DataFrame(
            [
                {"date": date(2026, 1, 1), "campaign_id": "1", "campaign_name": "A", "objective": "X",
                 "spend": 100.0, "impressions": 1000.0, "clicks": 100.0, "purchases": 1.0,
                 "purchase_value": 100.0, "add_to_carts": 5.0, "initiate_checkouts": 2.0, "frequency": 1.0},
                {"date": date(2026, 1, 2), "campaign_id": "1", "campaign_name": "A", "objective": "X",
                 "spend": 100.0, "impressions": 9000.0, "clicks": 90.0, "purchases": 1.0,
                 "purchase_value": 100.0, "add_to_carts": 5.0, "initiate_checkouts": 2.0, "frequency": 1.0},
            ]
        )
        totals = aggregate_campaign_totals(df)
        self.assertEqual(len(totals), 1)
        self.assertAlmostEqual(totals[0]["ctr"], 190.0 / 10000.0)


class TestAiManagementBriefSheet(ExcelReportTestCase):
    def test_all_required_sections_present(self):
        wb = self.generate()
        ws = wb["AI Management Brief"]
        all_text = [cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)]
        for label in (
            "Executive Verdict", "What Materially Changed", "Where Money Is Being Wasted",
            "Where Incremental Budget Should Move", "Growth Opportunities", "Funnel & Conversion Risks",
            "Tracking & Measurement Risks", "7-Day Action Plan", "Management Decisions Required",
        ):
            self.assertIn(label, all_text)

    def test_management_decisions_required_structure(self):
        wb = self.generate()
        ws = wb["AI Management Brief"]
        all_text = [cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)]
        for field in ("Priority:", "Decision:", "Evidence:", "Commercial Implication:", "Confidence:"):
            self.assertIn(field, all_text)


class TestAiEmptySectionFallback(unittest.TestCase):
    class _EmptyCase(ExcelReportTestCase):
        REPORT_MARKDOWN = EMPTY_SECTIONS_REPORT_MARKDOWN

    def setUp(self):
        self.case = self._EmptyCase()
        self.case.setUp()

    def tearDown(self):
        self.case.tearDown()

    def test_empty_sections_use_deterministic_fallback_not_forbidden_phrase(self):
        wb = self.case.generate()
        ws = wb["AI Management Brief"]
        all_text = " ".join(cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str))
        self.assertIn(NO_CONTENT_FALLBACK, all_text)
        self.assertNotIn(LEGACY_FORBIDDEN_PHRASE, all_text)


class TestExecutiveVerdictWordLimit(unittest.TestCase):
    def test_limit_words_truncates_over_100_words(self):
        text = " ".join(f"word{i}" for i in range(150))
        limited = limit_words(text, 100)
        self.assertEqual(len(limited.replace("…", "").split()), 100)
        self.assertTrue(limited.endswith("…"))

    def test_limit_words_leaves_short_text_untouched(self):
        text = "Short verdict text."
        self.assertEqual(limit_words(text, 100), text)


class TestInternalNavigation(ExcelReportTestCase):
    def test_back_to_cockpit_link_present_on_other_sheets(self):
        wb = self.generate()
        for sheet_name in ("Lifetime Performance", "Marketing Funnel", "Campaign Portfolio", "Trend & Efficiency"):
            ws = wb[sheet_name]
            hyperlinks = [cell.hyperlink for row in ws.iter_rows() for cell in row if cell.hyperlink]
            self.assertTrue(
                any("Executive Cockpit" in (h.target or "") for h in hyperlinks),
                f"No back-to-cockpit link found on '{sheet_name}'",
            )

    def test_cockpit_has_nav_links_to_other_sheets(self):
        wb = self.generate()
        ws = wb["Executive Cockpit"]
        hyperlinks = [cell.hyperlink for row in ws.iter_rows() for cell in row if cell.hyperlink]
        self.assertGreater(len(hyperlinks), 0)


class TestDecisionsRequiredStructure(ExcelReportTestCase):
    def test_cockpit_decisions_required_section_has_expected_columns(self):
        wb = self.generate()
        ws = wb["Executive Cockpit"]
        all_text = [cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)]
        self.assertIn("DECISIONS REQUIRED THIS WEEK", all_text)
        for col in ("Priority", "Decision", "Evidence", "Commercial Implication", "Confidence"):
            self.assertIn(col, all_text)


class TestAiAnalysisParsing(unittest.TestCase):
    def test_sections_split_correctly(self):
        sections = parse_ai_report_sections(SAMPLE_REPORT_MARKDOWN)
        self.assertIn("EXECUTIVE VERDICT", sections)
        self.assertIn("mixed", sections["EXECUTIVE VERDICT"])
        self.assertIn("Zero Purchase Campaign", sections["WHERE MONEY IS BEING WASTED"])
        self.assertEqual(
            sections["GROWTH OPPORTUNITIES"], "No issues detected in this category for the current period."
        )

    def test_recommendation_blocks_parsed_into_fields(self):
        sections = parse_ai_report_sections(SAMPLE_REPORT_MARKDOWN)
        blocks = parse_recommendation_blocks(sections["7-DAY ACTION PLAN"])
        self.assertEqual(len(blocks), 1)
        self.assertEqual(blocks[0]["Priority"], "critical")
        self.assertEqual(blocks[0]["Campaign"], "Zero Purchase Campaign")

    def test_malformed_block_falls_back_to_free_text_without_dropping_content(self):
        text = "Just some free text without field labels."
        blocks = parse_recommendation_blocks(text)
        self.assertEqual(len(blocks), 1)
        self.assertEqual(blocks[0]["free_text"], text)


if __name__ == "__main__":
    unittest.main()
