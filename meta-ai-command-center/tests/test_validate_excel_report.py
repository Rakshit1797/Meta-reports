#!/usr/bin/env python3
"""
Unit tests for validate_excel_report.py.

Builds a real workbook via excel_report.py from synthetic data, then checks
that validation passes on it and fails clearly on deliberately broken
variants. No live API calls are made. Run with:

    python -m unittest discover -s tests -t . -v
"""

import json
import os
import shutil
import sys
import tempfile
import unittest
from datetime import date, timedelta

import openpyxl
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from excel_report import generate_workbook  # noqa: E402
from validate_excel_report import ExcelValidationError, validate_workbook  # noqa: E402


def _build_synthetic_csv(path: str, include_lpv: bool = True) -> None:
    rows = []
    end = date(2026, 7, 14)
    start = end - timedelta(days=13)
    for offset in range(14):
        d = start + timedelta(days=offset)
        rows.append(
            {
                "date": d.isoformat(), "campaign_id": "400000000001", "campaign_name": "Campaign One",
                "objective": "OUTCOME_SALES", "spend": 100.0, "impressions": 3000.0, "reach": 2500.0,
                "clicks": 80.0, "ctr": 2.5, "cpc": 1.25, "cpm": 33.3, "frequency": 1.1,
                "purchases": 5.0, "purchase_value": 500.0, "CPA": 20.0, "ROAS": 5.0,
                "add_to_carts": 20.0, "initiate_checkouts": 10.0,
                "landing_page_views": 50.0 if include_lpv else None,
                "click_to_landing_page_rate": 0.625 if include_lpv else None,
                "landing_page_to_purchase_rate": 0.1 if include_lpv else None,
            }
        )
    pd.DataFrame(rows).to_csv(path, index=False)


SAMPLE_REPORT_MARKDOWN = """# Meta Ads Performance Intelligence Report

## EXECUTIVE VERDICT
Account performance was stable over the collected period with no material issues identified.

## WHAT MATERIALLY CHANGED
FACT: Metrics were broadly stable period over period.

## WHERE MONEY IS BEING WASTED
No issues detected in this category for the current period.

## WHERE INCREMENTAL BUDGET SHOULD MOVE
No issues detected in this category for the current period.

## GROWTH OPPORTUNITIES
No issues detected in this category for the current period.

## FUNNEL & CONVERSION RISKS
FACT: Funnel metrics were within normal range.

## TRACKING & MEASUREMENT RISKS
No issues detected in this category for the current period.

## 7-DAY ACTION PLAN
No issues detected in this category for the current period.

## MANAGEMENT DECISIONS REQUIRED
No issues detected in this category for the current period.
"""


class ValidateExcelReportTestCase(unittest.TestCase):
    INCLUDE_LPV = True

    def setUp(self):
        self.tmp_dir = tempfile.mkdtemp()
        self.csv_path = os.path.join(self.tmp_dir, "meta_campaign_daily.csv")
        self.findings_path = os.path.join(self.tmp_dir, "performance_findings.json")
        self.report_path = os.path.join(self.tmp_dir, "meta_performance_report.md")
        self.output_path = os.path.join(self.tmp_dir, "meta_ads_performance_intelligence.xlsx")

        _build_synthetic_csv(self.csv_path, include_lpv=self.INCLUDE_LPV)

        from performance_analyzer import analyze, load_dataset

        df = load_dataset(self.csv_path)
        results = analyze(df)
        with open(self.findings_path, "w") as f:
            json.dump(results, f, indent=2, default=str)
        with open(self.report_path, "w") as f:
            f.write(SAMPLE_REPORT_MARKDOWN)

        generate_workbook(self.csv_path, self.findings_path, self.report_path, self.output_path)

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def _account_currency(self):
        with open(self.findings_path) as f:
            return json.load(f).get("analysis_metadata", {}).get("account_currency")

    def _set_account_currency(self, currency_code):
        with open(self.findings_path) as f:
            findings = json.load(f)
        findings["analysis_metadata"]["account_currency"] = currency_code
        with open(self.findings_path, "w") as f:
            json.dump(findings, f, indent=2, default=str)
        generate_workbook(self.csv_path, self.findings_path, self.report_path, self.output_path)


class TestValidationSucceedsOnGoodWorkbook(ValidateExcelReportTestCase):
    def test_valid_workbook_passes(self):
        # Should not raise.
        validate_workbook(
            self.output_path, [self.csv_path, self.findings_path, self.report_path],
            account_currency=self._account_currency(), lpv_available=True,
        )


class TestValidationFailures(ValidateExcelReportTestCase):
    def test_missing_workbook_fails(self):
        with self.assertRaises(ExcelValidationError):
            validate_workbook(os.path.join(self.tmp_dir, "does_not_exist.xlsx"), [])

    def test_empty_workbook_file_fails(self):
        empty_path = os.path.join(self.tmp_dir, "empty.xlsx")
        open(empty_path, "w").close()
        with self.assertRaises(ExcelValidationError):
            validate_workbook(empty_path, [])

    def test_wrong_worksheet_order_fails(self):
        wb = openpyxl.load_workbook(self.output_path)
        wb.move_sheet("Raw Data", offset=-6)
        bad_path = os.path.join(self.tmp_dir, "wrong_order.xlsx")
        wb.save(bad_path)
        with self.assertRaises(ExcelValidationError):
            validate_workbook(bad_path, [])

    def test_missing_worksheet_fails(self):
        wb = openpyxl.load_workbook(self.output_path)
        del wb["AI Management Brief"]
        bad_path = os.path.join(self.tmp_dir, "missing_sheet.xlsx")
        wb.save(bad_path)
        with self.assertRaises(ExcelValidationError):
            validate_workbook(bad_path, [])

    def test_missing_companion_output_file_fails(self):
        with self.assertRaises(ExcelValidationError):
            validate_workbook(self.output_path, [os.path.join(self.tmp_dir, "does_not_exist.csv")])

    def test_empty_raw_data_sheet_fails(self):
        wb = openpyxl.load_workbook(self.output_path)
        ws = wb["Raw Data"]
        ws.delete_rows(2, ws.max_row)
        bad_path = os.path.join(self.tmp_dir, "empty_raw_data.xlsx")
        wb.save(bad_path)
        with self.assertRaises(ExcelValidationError):
            validate_workbook(bad_path, [])

    def test_missing_collected_period_dates_fails(self):
        wb = openpyxl.load_workbook(self.output_path)
        ws = wb["Lifetime Performance"]
        for row in ws.iter_rows(min_row=1, max_row=3):
            for cell in row:
                if isinstance(cell.value, str) and "Data available" in cell.value:
                    cell.value = "Redacted"
        bad_path = os.path.join(self.tmp_dir, "no_dates.xlsx")
        wb.save(bad_path)
        with self.assertRaises(ExcelValidationError):
            validate_workbook(bad_path, [])

    def test_ai_brief_empty_fails(self):
        wb = openpyxl.load_workbook(self.output_path)
        ws = wb["AI Management Brief"]
        # Strip the section body content (keep only the title/nav rows) so
        # the sheet is structurally empty of any required section label.
        if ws.max_row > 4:
            ws.delete_rows(5, ws.max_row - 4)
        bad_path = os.path.join(self.tmp_dir, "empty_ai_brief.xlsx")
        wb.save(bad_path)
        with self.assertRaises(ExcelValidationError):
            validate_workbook(bad_path, [])

    def test_ai_brief_forbidden_legacy_phrase_fails(self):
        wb = openpyxl.load_workbook(self.output_path)
        ws = wb["AI Management Brief"]
        ws.cell(row=ws.max_row + 2, column=1, value="No content generated for this section.")
        bad_path = os.path.join(self.tmp_dir, "forbidden_phrase.xlsx")
        wb.save(bad_path)
        with self.assertRaises(ExcelValidationError):
            validate_workbook(bad_path, [])

    def test_broken_chart_reference_fails(self):
        wb = openpyxl.load_workbook(self.output_path)
        ws = wb["Executive Cockpit"]
        broken = False
        for chart in ws._charts:
            for series in chart.series:
                if series.val is not None and series.val.numRef is not None:
                    series.val.numRef.f = ""
                    broken = True
        self.assertTrue(broken, "Expected at least one chart series to break for this test")
        bad_path = os.path.join(self.tmp_dir, "broken_chart.xlsx")
        wb.save(bad_path)
        with self.assertRaises(ExcelValidationError):
            validate_workbook(bad_path, [])

    def test_hardcoded_dollar_format_fails_when_account_currency_is_inr(self):
        self._set_account_currency("INR")
        wb = openpyxl.load_workbook(self.output_path)
        ws = wb["Raw Data"]
        header = [c.value for c in ws[1]]
        spend_col = header.index("spend") + 1
        ws.cell(row=2, column=spend_col).number_format = '"$"#,##0.00'
        bad_path = os.path.join(self.tmp_dir, "hardcoded_usd.xlsx")
        wb.save(bad_path)
        with self.assertRaises(ExcelValidationError) as ctx:
            validate_workbook(bad_path, [], account_currency="INR")
        self.assertIn("hardcoded", str(ctx.exception).lower())

    def test_inr_workbook_with_correct_symbol_passes_currency_check(self):
        self._set_account_currency("INR")
        with open(self.findings_path) as f:
            findings = json.load(f)
        validate_workbook(self.output_path, [], account_currency=findings["analysis_metadata"]["account_currency"])


class TestValidationLpvUnavailable(ValidateExcelReportTestCase):
    INCLUDE_LPV = False

    def test_lpv_shown_as_zero_when_unavailable_fails(self):
        wb = openpyxl.load_workbook(self.output_path)
        ws = wb["Executive Cockpit"]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "Website Landings (LPV)" in cell.value:
                    ws.cell(row=cell.row, column=cell.column + 1, value=0)
        bad_path = os.path.join(self.tmp_dir, "lpv_zero.xlsx")
        wb.save(bad_path)
        with self.assertRaises(ExcelValidationError) as ctx:
            validate_workbook(bad_path, [], lpv_available=False)
        self.assertIn("LPV", str(ctx.exception))

    def test_lpv_na_when_unavailable_passes(self):
        # generate_workbook itself must already render N/A, not 0, for LPV --
        # validation should pass cleanly against its own real output.
        validate_workbook(self.output_path, [], lpv_available=False)


if __name__ == "__main__":
    unittest.main()
