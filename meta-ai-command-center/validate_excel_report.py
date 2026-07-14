#!/usr/bin/env python3
"""
Validation for the Meta Ads Performance Intelligence Excel workbook.

Runs after excel_report.py and before email delivery. Every check here is
structural/content-presence validation -- it never recalculates or second-
guesses any metric, it only confirms the workbook was built as expected.

If validation fails, this script exits with code 1 and prints a clear,
human-readable error for each failed check. No email is sent when this
script fails (send_email_report.py is only invoked after this succeeds).
"""

import argparse
import os
import sys
from typing import List, Optional, Tuple

import openpyxl

from excel_report import (
    LEGACY_FORBIDDEN_PHRASE,
    NO_CONTENT_FALLBACK,
    REQUIRED_WORKSHEET_ORDER,
)

COCKPIT_CHART_TITLES = [
    "Spend vs Tracked Sales Trend",
    "ROAS Trend",
    "CPA Trend",
    "Purchases Trend",
    "Marketing Funnel (Last 7 Days)",
    "Top Campaign Spend Concentration",
    "Top Campaigns by Tracked Sales",
]
# Kept for backward compatibility with anything importing the old name.
REQUIRED_CHART_TITLES = COCKPIT_CHART_TITLES

DEFAULT_WORKBOOK_PATH = "meta_ads_performance_intelligence.xlsx"
DEFAULT_OTHER_OUTPUTS = [
    os.path.join("data", "meta_campaign_daily.csv"),
    os.path.join("data", "collection_metadata.json"),
    os.path.join("data", "performance_findings.json"),
    os.path.join("reports", "meta_performance_report.md"),
]


class ExcelValidationError(Exception):
    """Raised when the workbook (or a companion output file) fails validation."""


def _chart_title_text(chart) -> Optional[str]:
    try:
        return chart.title.tx.rich.p[0].r[0].t
    except Exception:
        return None


def _chart_series_references(chart) -> List[Optional[str]]:
    refs = []
    for series in chart.series:
        ref = None
        if series.val is not None and series.val.numRef is not None:
            ref = series.val.numRef.f
        elif getattr(series, "yVal", None) is not None and series.yVal.numRef is not None:
            ref = series.yVal.numRef.f
        refs.append(ref)
    return refs


def check_file_exists(path: str) -> Tuple[bool, str]:
    if not os.path.isfile(path):
        return False, f"Required output file '{path}' does not exist."
    return True, ""


def check_file_not_empty(path: str) -> Tuple[bool, str]:
    if os.path.getsize(path) == 0:
        return False, f"Required output file '{path}' is empty."
    return True, ""


def check_workbook_opens(path: str) -> Tuple[bool, str, Optional[openpyxl.Workbook]]:
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as exc:
        return False, f"Workbook '{path}' could not be opened with openpyxl: {exc}", None
    return True, "", wb


def check_worksheet_set_and_order(wb: openpyxl.Workbook) -> List[str]:
    errors = []
    actual = list(wb.sheetnames)
    if set(actual) != set(REQUIRED_WORKSHEET_ORDER):
        errors.append(
            f"Workbook worksheets {actual} do not match the required set {REQUIRED_WORKSHEET_ORDER}."
        )
    elif actual != REQUIRED_WORKSHEET_ORDER:
        errors.append(
            f"Worksheet order is {actual}, but must be exactly {REQUIRED_WORKSHEET_ORDER}."
        )
    return errors


def check_raw_data_sheet(wb: openpyxl.Workbook) -> List[str]:
    errors = []
    if "Raw Data" not in wb.sheetnames:
        return ["'Raw Data' worksheet is missing."]
    ws = wb["Raw Data"]
    if ws.max_row < 2:
        errors.append("'Raw Data' worksheet contains no data rows (only a header, or is empty).")
    return errors


def check_executive_cockpit_sheet(wb: openpyxl.Workbook) -> List[str]:
    errors = []
    if "Executive Cockpit" not in wb.sheetnames:
        return ["'Executive Cockpit' worksheet is missing."]
    ws = wb["Executive Cockpit"]

    kpi_values_found = False
    for row in ws.iter_rows(min_row=1, max_row=25, max_col=6):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                kpi_values_found = True
                break
        if kpi_values_found:
            break
    if not kpi_values_found:
        errors.append("'Executive Cockpit' worksheet has no numeric KPI card values.")

    found_titles = {_chart_title_text(chart) for chart in ws._charts}
    missing_charts = [title for title in COCKPIT_CHART_TITLES if title not in found_titles]
    if missing_charts:
        errors.append(f"'Executive Cockpit' is missing required chart(s): {', '.join(missing_charts)}.")

    for chart in ws._charts:
        title = _chart_title_text(chart) or "<untitled chart>"
        refs = _chart_series_references(chart)
        if not refs:
            errors.append(f"Chart '{title}' has no data series.")
            continue
        for ref in refs:
            # A broken/blanked reference can round-trip through openpyxl's
            # XML serialization as the literal string "None" rather than a
            # real empty value -- treat that the same as missing.
            if not ref or ref == "None":
                errors.append(f"Chart '{title}' has a series with no worksheet cell reference (broken source).")

    return errors


def check_collected_period_sheet(wb: openpyxl.Workbook) -> List[str]:
    errors = []
    if "Lifetime Performance" not in wb.sheetnames:
        return ["'Lifetime Performance' worksheet is missing."]
    ws = wb["Lifetime Performance"]

    dates_found = False
    for row in ws.iter_rows(min_row=1, max_row=3, max_col=8):
        for cell in row:
            if isinstance(cell.value, str) and "Data available" in cell.value:
                dates_found = True
    if not dates_found:
        errors.append(
            "'Lifetime Performance' worksheet is missing the collected-period date disclosure "
            "('Data available from ... to ...')."
        )
    return errors


def check_campaign_portfolio_sheet(wb: openpyxl.Workbook) -> List[str]:
    errors = []
    if "Campaign Portfolio" not in wb.sheetnames:
        return ["'Campaign Portfolio' worksheet is missing."]
    ws = wb["Campaign Portfolio"]
    if ws.max_row < 6:
        errors.append("'Campaign Portfolio' worksheet contains no campaign rows.")
    return errors


def check_ai_management_brief_sheet(wb: openpyxl.Workbook) -> List[str]:
    from excel_report import REQUIRED_AI_BRIEF_SECTIONS, AI_BRIEF_SECTION_LABELS

    errors = []
    if "AI Management Brief" not in wb.sheetnames:
        return ["'AI Management Brief' worksheet is missing."]
    ws = wb["AI Management Brief"]

    all_values = [cell.value for row in ws.iter_rows(max_col=2) for cell in row]
    forbidden_phrase_found = any(
        isinstance(v, str) and LEGACY_FORBIDDEN_PHRASE in v for v in all_values
    )

    # A genuinely populated brief must contain every required section label
    # (Executive Verdict, What Materially Changed, ...) -- not just any
    # string cell, since decoration (title/nav links) alone must not count
    # as "report content".
    required_labels = set(AI_BRIEF_SECTION_LABELS.values())
    labels_present = {v for v in all_values if isinstance(v, str)} & required_labels
    missing_labels = required_labels - labels_present
    if missing_labels:
        errors.append(
            f"'AI Management Brief' worksheet contains no report content (missing section(s): "
            f"{', '.join(sorted(missing_labels))})."
        )
    if forbidden_phrase_found:
        errors.append(
            f"'AI Management Brief' worksheet contains the forbidden placeholder phrase "
            f"'{LEGACY_FORBIDDEN_PHRASE}' -- use '{NO_CONTENT_FALLBACK}' instead."
        )
    return errors


def check_no_hardcoded_usd_for_non_usd_currency(wb: openpyxl.Workbook, account_currency: Optional[str]) -> List[str]:
    """Fail if any currency-formatted cell uses a literal '$' number_format while the
    account's real currency is a known non-USD currency (e.g. INR)."""
    if not account_currency or account_currency.upper() == "USD":
        return []
    errors = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                fmt = cell.number_format or ""
                if "$" in fmt:
                    errors.append(
                        f"Sheet '{sheet_name}' cell {cell.coordinate} uses a hardcoded '$' number "
                        f"format, but the account currency is '{account_currency}'."
                    )
    return errors


def check_lpv_never_zero_when_unavailable(wb: openpyxl.Workbook, lpv_available: bool) -> List[str]:
    """Fail if LPV is unavailable in the source data but any LPV-labeled cell shows 0
    instead of N/A (a fabricated zero, not a genuine absence indicator)."""
    if lpv_available:
        return []
    errors = []
    lpv_labels = ("LPV", "Website Landings")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(label in cell.value for label in lpv_labels):
                    # Look one cell to the right for a value cell that might hold a
                    # fabricated zero where N/A is required.
                    neighbor = ws.cell(row=cell.row, column=cell.column + 1)
                    if neighbor.value == 0:
                        errors.append(
                            f"Sheet '{sheet_name}' cell {neighbor.coordinate} shows 0 for an LPV-labeled "
                            f"metric, but LPV is unavailable in the source data -- must be N/A, not 0."
                        )
    return errors


def validate_workbook(
    workbook_path: str, other_output_paths: List[str], account_currency: Optional[str] = None, lpv_available: bool = True
) -> None:
    """Run every validation check. Raises ExcelValidationError with all failures if any fail."""
    errors: List[str] = []

    ok, message = check_file_exists(workbook_path)
    if not ok:
        raise ExcelValidationError(message)
    ok, message = check_file_not_empty(workbook_path)
    if not ok:
        raise ExcelValidationError(message)

    for path in other_output_paths:
        ok, message = check_file_exists(path)
        if not ok:
            errors.append(message)
            continue
        ok, message = check_file_not_empty(path)
        if not ok:
            errors.append(message)

    ok, message, wb = check_workbook_opens(workbook_path)
    if not ok:
        raise ExcelValidationError(message)

    errors.extend(check_worksheet_set_and_order(wb))
    errors.extend(check_raw_data_sheet(wb))
    errors.extend(check_executive_cockpit_sheet(wb))
    errors.extend(check_collected_period_sheet(wb))
    errors.extend(check_campaign_portfolio_sheet(wb))
    errors.extend(check_ai_management_brief_sheet(wb))
    errors.extend(check_no_hardcoded_usd_for_non_usd_currency(wb, account_currency))
    errors.extend(check_lpv_never_zero_when_unavailable(wb, lpv_available))

    if errors:
        raise ExcelValidationError("\n".join(f"  - {err}" for err in errors))


def _load_account_currency_and_lpv(findings_path: str) -> Tuple[Optional[str], bool]:
    import json

    try:
        with open(findings_path) as f:
            findings = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return None, True

    account_currency = findings.get("analysis_metadata", {}).get("account_currency")
    last7 = findings.get("account_summary", {}).get("last_7_days", {})
    lpv_available = last7.get("landing_page_views") is not None
    return account_currency, lpv_available


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate the Meta Ads Performance Intelligence Excel workbook.")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK_PATH, help="Path to the .xlsx workbook to validate.")
    parser.add_argument(
        "--other-outputs",
        nargs="*",
        default=DEFAULT_OTHER_OUTPUTS,
        help="Other pipeline output files that must exist and be non-empty.",
    )
    parser.add_argument(
        "--findings",
        default=os.path.join("data", "performance_findings.json"),
        help="Path to performance_findings.json, used to read the account currency and LPV availability.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    account_currency, lpv_available = _load_account_currency_and_lpv(args.findings)
    try:
        validate_workbook(args.workbook, args.other_outputs, account_currency, lpv_available)
    except ExcelValidationError as exc:
        print("Excel workbook validation FAILED:")
        print(str(exc))
        sys.exit(1)

    print(f"Excel workbook validation passed: '{args.workbook}' is well-formed.")


if __name__ == "__main__":
    main()
